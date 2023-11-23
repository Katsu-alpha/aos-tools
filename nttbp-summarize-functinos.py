#
#   以下の設定を SSID ごとにサマライズ
#
#   MAC認証 (aaa profile)
#   1x 再認証 (aaa authentication dot1x)
#   delete-keycache (aaa authentication dot1x)
#   hide-ssid (ssid profile)
#   deny-inter-user-traffic (virtual-ap profile)
#   W52/W53/W56/custom (regulatory-domain profile)
#   XML-API (aaa profile)
#   Radius accounting
#   Radius interim accounting
#   ClientMatch (ARM profile)
#   ClientMatch custom (ARM profile)
#   ARM dynamic channel (ARM profile)
#   Static channel (ARM profile)
#   Local probe req threshold (ssid profile)
#   Auth req threshold (ssid profile)
#   VHT disable (ht-ssid profile)
#   HT disable (ht-ssid profile)
#

import re
import glob
import sys
import argparse
import mylogger as log
from aos_parser import AOSParser
from colorama import Fore, Style
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict


def print_debug(msg):
    print(Fore.GREEN + msg + Style.RESET_ALL)

def print_info(msg):
    print(Fore.CYAN + msg + Style.RESET_ALL)

def print_warn(msg):
    print(Fore.YELLOW + msg + Style.RESET_ALL)

def print_err(msg):
    print(Fore.RED + msg + Style.RESET_ALL)


class SSID:
    def __init__(self, name):
        self.name = name
        self.ssid = ''
        self.opmode = 'open'
        self.stealth = False
        self.disabled = False
        self.htssid_prof = None
        self.probe_thresh = False
        self.auth_thresh = False
        self.adv_apname = False

class VAP:
    def __init__(self, name):
        self.name = name
        self.ssid_prof = None
        self.aaa_prof = None
        self.fwd_mode = 'tunnel'
        self.vlan = 1
        self.disabled = False
        self.deny_int_user = False
        self.bcast_filter_all = False
        self.cha = False

class APGroup:
    def __init__(self, name):
        self.name = name
        self.vaps = []
        self.dot11a = None
        self.dot11g = None
        self.reg_prof = None

class APName:
    def __init__(self, name):
        self.name = name
        self.vaps = []
        self.exclude_vaps = []
        self.dot11a = None
        self.dot11g = None
        self.reg_prof = None
        self.misc = []

class ARM:
    def __init__(self, name):
        self.name = name
        self.disabled = False
        self.cm = True
        self.cm_custom = False

class DOT11A:
    def __init__(self, name):
        self.name = name
        self.arm_prof = None
        self.radio = True

class DOT11G:
    def __init__(self, name):
        self.name = name
        self.arm_prof = None
        self.radio = True


class AAA:
    def __init__(self, name):
        self.name = name
        self.mac = False
        self.dot1x = False
        self.init_role = 'logon'
        self.xmlapi = False
        self.acct = False
        self.interim_acct = False
        self.user_deriv = False
        self.l2_fail_thru = False

class HTSSID:
    def __init__(self, name):
        self.name = name
        self.ht = True
        self.vht = True
        self.eighty = True
        self.forty = True

class REGDOM:
    def __init__(self, name):
        self.name = name
        self.chset_11a = ""
        self.band = None

class DOT1X:
    def __init__(self, name):
        self.name = name
        self.reauth = False
        self.delete_keycache = False

class AOSConfig:
    def __init__(self, name):
        self.name = name        # ホスト名
        self.num_user_roles = 0 # user-role 数
        self.num_acls = 0       # session ACL 数
        self.ssid_profs = {}    # key=プロファイル名, val=SSID obj
        self.vap_profs = {}     # key=プロファイル名, val=VAP obj
        self.ap_groups = {}     # key=AP グループ名, val=APGroup obj
        self.ap_names = {}      # key=AP 名, val=APName obj
        self.arm_profs = {}
        self.dot11a_profs = {}
        self.dot11g_profs = {}
        self.aaa_profs = {}
        self.htssid_profs = {}
        self.regdom_profs = {}
        self.dot1x_profs = {}


def parse_config(name, config, conf_ver):
    C = AOSConfig(name)

    dot11a_def = DOT11A('default')
    C.dot11a_profs['default'] = dot11a_def
    dot11g_def = DOT11G('default')
    C.dot11g_profs['default'] = dot11g_def
    regdom_def = REGDOM('default')
    regdom_def.band = 'W52/W53/W56'
    C.regdom_profs['default'] = regdom_def
    dot1x_def = DOT1X('default')
    C.dot1x_profs['default'] = dot1x_def

    if conf_ver == 8:
        arm_def_a = ARM('default-a')
        C.arm_profs['default-a'] = arm_def_a
        arm_def_g = ARM('default-g')
        C.arm_profs['default-g'] = arm_def_g

    in_cont = False

    for l in config:
        if not in_cont:

            if l.startswith('ip access-list session '):
                C.num_acls += 1  # ip access-list 数カウント
                continue

            if l.startswith('user-role '):
                C.num_user_roles += 1  # user-role 数カウント
                continue

            if l.startswith('ap system-profile '):
                in_cont = 'apsys'
                name = l[18:].replace('"', '')
                continue

            if l.startswith('wlan ssid-profile '):
                in_cont = 'ssid'
                name = l[18:].replace('"', '')
                ssid = SSID(name)
                C.ssid_profs[name] = ssid
                continue

            if l.startswith('wlan virtual-ap '):
                in_cont = 'vap'
                name = l[16:].replace('"', '')
                vap = VAP(name)
                C.vap_profs[name] = vap
                continue

            if l.startswith('ap-group '):
                in_cont = 'apg'
                name = l[9:].replace('"', '').lower()
                apg = APGroup(name)
                C.ap_groups[name] = apg
                continue

            if l.startswith('ap-name '):
                in_cont = 'apn'
                name = l[8:].replace('"', '').lower()
                apn = APName(name)
                continue

            if l.startswith('rf arm-profile '):
                in_cont = 'arm'
                name = l[15:].replace('"', '')
                arm = ARM(name)
                C.arm_profs[name] = arm
                continue

            if l.startswith('rf dot11a-radio-profile '):
                in_cont = 'dot11a'
                name = l[24:].replace('"', '')
                dot11a = DOT11A(name)
                dot11a.arm_prof = C.arm_profs['default-a']
                C.dot11a_profs[name] = dot11a
                continue

            if l.startswith('rf dot11g-radio-profile '):
                in_cont = 'dot11g'
                name = l[24:].replace('"', '')
                dot11g = DOT11G(name)
                dot11g.arm_prof = C.arm_profs['default-g']
                C.dot11g_profs[name] = dot11g
                continue

            if l.startswith('aaa profile '):
                in_cont = 'aaa'
                name = l[12:].replace('"', '')
                aaa = AAA(name)
                C.aaa_profs[name] = aaa
                continue

            if l.startswith('wlan ht-ssid-profile '):
                in_cont = 'htssid'
                name = l[21:].replace('"', '')
                htssid = HTSSID(name)
                C.htssid_profs[name] = htssid
                continue

            if l.startswith('ap regulatory-domain-profile '):
                in_cont = 'regdom'
                name = l[29:].replace('"', '')
                regdom = REGDOM(name)
                C.regdom_profs[name] = regdom
                continue

            if l.startswith('aaa authentication dot1x '):
                in_cont = 'dot1x'
                name = l[25:].replace('"', '')
                dot1x = DOT1X(name)
                C.dot1x_profs[name] = dot1x
                continue

            continue

        #
        #   inside a context
        #

        l = l.strip()
        if l == '!':
            if in_cont == 'regdom':
                if regdom.chset_11a == '36,40,44,48,36-40,44-48,':
                    regdom.band = 'W52'
                elif regdom.chset_11a == '52,56,60,64,52-56,60-64,':
                    regdom.band = 'W53'
                elif regdom.chset_11a == '100,104,108,112,116,120,124,128,132,136,140,100-104,108-112,116-120,124-128,132-136,':
                    regdom.band = 'W56'
                elif regdom.chset_11a == '36,40,44,48,52,56,60,64,36-40,44-48,52-56,60-64,':
                    regdom.band = 'W52/W53'
                elif regdom.chset_11a == '36,40,44,48,100,104,108,112,116,120,124,128,132,136,140,36-40,44-48,100-104,108-112,116-120,124-128,132-136,':
                    regdom.band = 'W52/W56'
                elif regdom.chset_11a == '52,56,60,64,100,104,108,112,116,120,124,128,132,136,140,52-56,60-64,100-104,108-112,116-120,124-128,132-136,':
                    regdom.band = 'W53/W56'
                elif regdom.chset_11a == '36,40,44,48,52,56,60,64,100,104,108,112,116,120,124,128,132,136,140,36-40,44-48,52-56,60-64,100-104,108-112,116-120,124-128,132-136,':
                    regdom.band = 'W52/W53/W56'
                else:
                    regdom.band = 'Custom'

            in_cont = False
            continue

        if in_cont == 'apsys':
            r = re.match(r'telnet|dns-domain |lms-ip |bkup-lms-ip |bootstrap-threshold |number_ipsec_retries |secondary-master |shell-passwd |bkup-passwords |lms-preemption|ap-console-password |heartbeat-interval |session-acl ', l)
            if not r:
                print_warn(f'ap system-profile {name}: {l}')   # 特殊な system-profile 設定
                continue
            continue

        if in_cont == 'ssid':
            if l.startswith('essid '):
                ssid.ssid = l[6:].replace('"', '')
                continue
            if l.startswith('opmode '):
                ssid.opmode = l[7:]
                continue
            if l.startswith('hide-ssid'):
                ssid.stealth = True
                continue
            if l.startswith('no ssid-enable'):
                ssid.disabled = True
                continue
            if l.startswith('ht-ssid-profile '):
                n = l[16:].replace('"', '')
                ssid.htssid_prof = C.htssid_profs[n]
                continue
            if l.startswith('local-probe-req-thresh '):
                ssid.probe_thresh = True
                continue
            if l.startswith('auth-req-thresh '):
                ssid.auth_thresh = True
                continue
            if l.startswith('advertise-ap-name'):
                ssid.adv_apname = True
                continue
            r = re.match(r'dtim-period |wepkey1 |wpa-passphrase |deny-bcast|a-tx-rates |g-tx-rates |a-basic-rates |g-basic-rates |ageout |max-clients |no wmm-uapsd|wmm-|g-beacon-rate |a-beacon-rate ', l)
            if not r:
                print_warn(f'ssid-profile {name}: {l}')
            continue
            continue

        if in_cont == 'vap':
            if l.startswith('aaa-profile '):
                n = l[12:].replace('"', '')
                vap.aaa_prof = C.aaa_profs[n]
                continue
            if l.startswith('ssid-profile '):
                n = l[13:].replace('"', '')
                vap.ssid_prof = C.ssid_profs[n]
                if vap.ssid_prof.disabled:
                    vap.disabled = True
                continue
            if l.startswith('vlan '):
                vap.vlan = int(l[5:])
                continue
            if l.startswith('forward-mode '):
                vap.fwd_mode = l[13:]
                continue
            if l.startswith('deny-inter-user-traffic'):
                vap.deny_int_user = True
                continue
            if l.startswith('broadcast-filter all'):
                vap.bcast_filter_all = True
                continue
            if l.startswith('cellular-handoff-assist'):
                vap.cha = True
                continue

            r = re.match(r'allowed-band ', l)
            if not r:
                print_warn(f'virtual-ap {name}: {l}')
            continue

        if in_cont == 'apg':
            if l.startswith('virtual-ap '):
                n = l[11:].replace('"', '')
                apg.vaps.append(C.vap_profs[n])
                C.ap_groups[name] = apg
                continue
            if l.startswith('dot11a-radio-profile '):
                n = l[21:].replace('"', '')
                apg.dot11a = C.dot11a_profs[n]
                continue
            if l.startswith('dot11g-radio-profile '):
                n = l[21:].replace('"', '')
                apg.dot11g = C.dot11g_profs[n]
                continue
            if l.startswith('regulatory-domain-profile '):
                n = l[26:].replace('"', '')
                apg.reg_prof = C.regdom_profs[n]
                continue
            continue

        if in_cont == 'apn':
            if l.startswith('virtual-ap '):
                n = l[11:].replace('"', '')
                apn.vaps.append(C.vap_profs[n])
                C.ap_names[name] = apn
                continue
            if l.startswith('exclude-virtual-ap '):
                n = l[19:].replace('"', '')
                apn.exclude_vaps.append(n)
                C.ap_names[name] = apn
                continue
            if l.startswith('dot11a-radio-profile '):
                n = l[21:].replace('"', '')
                apn.dot11a = C.dot11a_profs[n]
                continue
            if l.startswith('dot11g-radio-profile '):
                n = l[21:].replace('"', '')
                apn.dot11g = C.dot11g_profs[n]
                continue
            if l.startswith('regulatory-domain-profile '):
                n = l[26:].replace('"', '')
                apn.reg_prof = C.regdom_profs[n]
                continue
            r = re.match(r'ap-system-profile ', l)
            if not r:
                print_warn(f'ap-name {name}: {l}')
            apn.misc.append(l)
            continue

        if in_cont == 'arm':
            if l.startswith('assignment disable'):
                arm.disabled = True
                continue
            if l.startswith('no client-match'):
                arm.cm = False
                continue
            if l.startswith('cm-'):
                arm.cm_custom = True
                continue
            continue

        if in_cont == 'dot11a':
            if l.startswith('no radio-enable'):
                dot11a.radio = False
                continue
            if l.startswith('arm-profile '):
                n = l[12:].replace('"', '')
                try:
                    dot11a.arm_prof = C.arm_profs[n]
                except KeyError:
                    print_err(f"In dot11a radio profile '{name}'")
                    print_err(f"ARM profile '{n}' referenced but not defined.")
                continue
            continue

        if in_cont == 'dot11g':
            if l.startswith('no radio-enable'):
                dot11g.radio = False
                continue
            if l.startswith('arm-profile '):
                n = l[12:].replace('"', '')
                try:
                    dot11g.arm_prof = C.arm_profs[n]
                except KeyError:
                    print_err(f"In dot11g radio profile '{name}'")
                    print_err(f"ARM profile '{n}' referenced but not defined.")
                continue
            continue

        if in_cont == 'aaa':
            if l.startswith('initial-role '):
                n = l[13:].replace('"', '')
                aaa.init_role = n
                continue
            if l.startswith('authentication-dot1x '):
                n = l[21:].replace('"', '')
                aaa.dot1x = C.dot1x_profs[n]
                continue
            if l.startswith('authentication-mac '):
                aaa.mac = True
                continue
            if l.startswith('xml-api-server '):
                aaa.xmlapi = True
                continue
            if l.startswith('radius-accounting '):
                aaa.acct = True
                continue
            if l.startswith('radius-interim-accounting'):
                aaa.interim_acct = True
                continue
            if l.startswith('user-derivation-rules '):
                aaa.user_deriv = True
                continue
            if l.startswith('l2-auth-fail-through'):
                aaa.l2_fail_thru = True
                continue

            r = re.match(r'no devtype-classification|mac-default-role |mac-server-group |dot1x-default-role |dot1x-server-group |user-idle-timeout ', l)
            if not r:
                print_warn(f'aaa profile {name}: {l}')
            continue

        if in_cont == 'htssid':
            if l.startswith('no high-throughput-enable'):
                htssid.ht = False
                continue
            if l.startswith('no very-high-throughput-enable'):
                htssid.vht = False
                continue
            if l.startswith('no 40MHz-enable'):
                htssid.forty = False
                continue
            if l.startswith('no 80MHz-enable'):
                htssid.eighty = False
                continue

            r = re.match(r'no ba-amsdu-enable|supported-mcs-set ', l)
            if not r:
                print_warn(f'ht ssid-profile {name}: {l}')
            continue

        if in_cont == 'regdom':
            if l.startswith('valid-11a-channel '):
                c = l[18:]
                regdom.chset_11a += c+','
                continue
            if l.startswith('valid-11a-40mhz-channel-pair '):
                c = l[29:]
                regdom.chset_11a += c+','
                continue
            continue

        if in_cont == 'dot1x':
            if l.startswith('reauthentication'):
                dot1x.reauth = True
                continue
            if l.startswith('delete-keycache'):
                dot1x.delete_keycache = True
                continue
            continue

    return C    # return an AOSConfig object

#-------------------------------------------------------------------------

#
#   start
#

parser = argparse.ArgumentParser(
    description="parse show run/show ap database and summarize")
parser.add_argument('files', help="show run/show ap database long outputs", type=str, nargs='*')
parser.add_argument('--debug', help='debug log', action='store_true')
args = parser.parse_args()

if args.debug:
    log.setloglevel(log.LOG_DEBUG)
else:
    log.setloglevel(log.LOG_INFO)

AP_DATABASE_LONG = "show ap database long"

if '*' in args.files[0]:
    args.files = glob.glob(args.files[0])

xlsfile = 'summary-func.xlsx'

#
#   process each file
#
global_vap_count = defaultdict(lambda: 0)       # key=VAP tuple


vap_stealth = defaultdict(lambda: 0)
vap_reauth = defaultdict(lambda: 0)
vap_macauth = defaultdict(lambda: 0)
vap_del_keycache = defaultdict(lambda: 0)
vap_deny_int_user = defaultdict(lambda: 0)
vap_bcast_filter = defaultdict(lambda: 0)
vap_xmlapi = defaultdict(lambda: 0)
vap_acct = defaultdict(lambda: 0)
vap_interim_acct = defaultdict(lambda: 0)
vap_user_deriv = defaultdict(lambda: 0)
vap_probe_thresh = defaultdict(lambda: 0)
vap_auth_thresh = defaultdict(lambda: 0)
vap_adv_apname = defaultdict(lambda: 0)
vap_vht_disable = defaultdict(lambda: 0)
vap_ht_disable = defaultdict(lambda: 0)
vap_no_bonding = defaultdict(lambda: 0)
vap_l2_fail_thru = defaultdict(lambda: 0)
vap_cha = defaultdict(lambda: 0)

vap_w52 = defaultdict(lambda: 0)
vap_w53 = defaultdict(lambda: 0)
vap_w56 = defaultdict(lambda: 0)
vap_w5253 = defaultdict(lambda: 0)
vap_w5256 = defaultdict(lambda: 0)
vap_w5356 = defaultdict(lambda: 0)
vap_reg_all = defaultdict(lambda: 0)
vap_reg_custom = defaultdict(lambda: 0)

vap_11a_off = defaultdict(lambda: 0)
vap_11g_off = defaultdict(lambda: 0)
vap_cm = defaultdict(lambda: 0)
vap_cm_custom = defaultdict(lambda: 0)
vap_static = defaultdict(lambda: 0)


for fn in args.files:
    print(f"Processing file: {fn}")

    # Parse 'show ap database long' table
    aos = AOSParser(fn, cmds=[AP_DATABASE_LONG])

    ap_db_tbl = aos.get_table(AP_DATABASE_LONG)
    if ap_db_tbl is None:
        print_info(f"No active AP found in {fn}")
        continue

    ap_db_tbl = ap_db_tbl[1:]   # discard header


    #
    #   Parse running-config, get controller model, version
    #
    in_showrun = False
    config = []
    with open(fn) as f:
        for line in f:
            if in_showrun:
                if re.match(r'\([\w\-_]+\) #', line):   # AOS6 prompt
                    break
                if re.match(r'\([\w\-_]+\) \*?\[[\w/]+\] #', line):   # AOS8 prompt
                    break
                config.append(line.rstrip())
                continue
            if re.search("show running-config", line):
                in_showrun = True
                conf_ver = 6
                continue
            if re.search(r'show configuration committed /mm/mynode', line):
                in_showrun = True
                conf_ver = 8
                continue
            r = re.match(r'ArubaOS \(MODEL: (\w+)\), Version ([\d\.]+)', line)
            if r:
                model = r.group(1)
                version = r.group(2)
                continue

    controller = fn[:-4]
    c = parse_config(controller, config, conf_ver)


    #   check static assignment, count used SSIDs
    vap_count = {}      # key=(ssid, opmode, fwd-mode)

    for r in ap_db_tbl:
        apg = r[1].lower()
        apn = r[0].lower()

        #   ARM profile/Radio profile
        armprof_a = None
        armprof_g = None

        dot11a = c.ap_groups[apg].dot11a
        dot11g = c.ap_groups[apg].dot11g
        regdom = c.ap_groups[apg].reg_prof

        both_radio_disabled = False

        #   AP-Name override
        if apn in c.ap_names:
            apn_c = c.ap_names[apn]
            if apn_c.dot11a:
                dot11a = apn_c.dot11a       # overrode by ap-name
            if apn_c.dot11g:
                dot11g = apn_c.dot11g
            if dot11a and dot11a.radio is False and dot11g and dot11g.radio is False:
                print_err(f"Both radios are disabled on {r[0]}!")
                both_radio_disabled = True

            if apn_c.reg_prof:
                regdom = apn_c.reg_prof

        if dot11a:
            armprof_a = dot11a.arm_prof
        if dot11g:
            armprof_g = dot11g.arm_prof

        ap_11a_off = False
        ap_11g_off = False
        ap_static = False
        if dot11a and dot11a.radio is False:
            ap_11a_off = True
        elif armprof_a and armprof_a.disabled:
            ap_static = True

        if dot11g and dot11g.radio is False:
            ap_11g_off = True
        elif armprof_g and armprof_g.disabled:
            ap_static = True

        if both_radio_disabled:
            continue

        if regdom is None:
            #print_warn(f'no regulatory domain defined in ap-group {apg} ap-name {apn}')
            ap_11a_band = 'W52/W53/W56'
        else:
            ap_11a_band = regdom.band

        #   ClientMatch
        ap_cm = False
        ap_cm_custom = False
        if armprof_a and armprof_a.cm:
            ap_cm = True
            if armprof_a.cm_custom:
                ap_cm_custom = True

        #   create VAP set for the AP
        vap_prof_set = set()
        for vap in c.ap_groups[apg].vaps:
            if not vap.disabled:
                vap_prof_set.add(vap)

        if apn in c.ap_names:
            apn_c = c.ap_names[apn]
            for vap_n in apn_c.exclude_vaps:
                vap = c.vap_profs[vap_n]
                vap_prof_set.discard(vap)   # will not raise exception
            for vap in apn_c.vaps:
                if not vap.disabled:
                    vap_prof_set.add(vap)

        #   create function sets for each VAP Tuple
        vap_tuple_set = set()
        for vap in vap_prof_set:
            vap_tuple = (vap.ssid_prof.ssid, vap.ssid_prof.opmode, vap.fwd_mode)
            if vap_tuple in vap_tuple_set:
                continue        # duplicate tuple (same for 2.4GHz/5GHz)

            vap_tuple_set.add(vap_tuple)
            global_vap_count[vap_tuple] += 1

            # if vap_tuple[0] == 'PremiumOutletsJP' and vap_tuple[2] == 'tunnel':
            #     print_err(f'PremiumOutlets: AP={apn}, group={apg}')


            #   VAP profile
            if vap.deny_int_user:
                vap_deny_int_user[vap_tuple] += 1
            if vap.cha:
                vap_cha[vap_tuple] += 1
            if vap.bcast_filter_all:
                vap_bcast_filter[vap_tuple] += 1

            #   SSID profile
            ssid = vap.ssid_prof
            if ssid.stealth:
                vap_stealth[vap_tuple] += 1
            if ssid.probe_thresh:
                vap_probe_thresh[vap_tuple] += 1
            if ssid.auth_thresh:
                vap_auth_thresh[vap_tuple] += 1
            if ssid.adv_apname:
                vap_adv_apname[vap_tuple] += 1

            #   HT-SSID profile
            htssid = ssid.htssid_prof
            if htssid:
                if htssid.ht is False:
                    vap_ht_disable[vap_tuple] += 1
                elif htssid.vht is False:
                    vap_vht_disable[vap_tuple] += 1
                if htssid.eighty is False and htssid.forty is False:
                    vap_no_bonding[vap_tuple] += 1

            #   AAA profile
            aaa = vap.aaa_prof
            if aaa.mac:
                vap_macauth[vap_tuple] += 1
            if aaa.xmlapi:
                vap_xmlapi[vap_tuple] += 1
            if aaa.acct:
                vap_acct[vap_tuple] += 1
            if aaa.interim_acct:
                vap_interim_acct[vap_tuple] += 1
            if aaa.user_deriv:
                vap_user_deriv[vap_tuple] += 1
            if aaa.l2_fail_thru:
                vap_l2_fail_thru[vap_tuple] += 1

            #   Dot1x profile
            dot1x = vap.aaa_prof.dot1x
            if dot1x:
                if dot1x.reauth:
                    vap_reauth[vap_tuple] += 1
                if dot1x.delete_keycache:
                    vap_del_keycache[vap_tuple] += 1

            #   11a band
            if ap_11a_band == 'W52':
                vap_w52[vap_tuple] += 1
            elif ap_11a_band == 'W53':
                vap_w53[vap_tuple] += 1
            elif ap_11a_band == 'W56':
                vap_w56[vap_tuple] += 1
            elif ap_11a_band == 'W52/W53':
                vap_w5253[vap_tuple] += 1
            elif ap_11a_band == 'W52/W56':
                vap_w5256[vap_tuple] += 1
            elif ap_11a_band == 'W53/W56':
                vap_w5356[vap_tuple] += 1
            elif ap_11a_band == 'W52/W53/W56':
                vap_reg_all[vap_tuple] += 1
            elif ap_11a_band == 'Custom':
                vap_reg_custom[vap_tuple] += 1
            else:
                print_err(f"Unknown 11a band '{ap_11a_band}' for AP {apn}")
                exit()

            #   Radio off
            if ap_11a_off:
                vap_11a_off[vap_tuple] += 1
            if ap_11g_off:
                vap_11g_off[vap_tuple] += 1

            #   Static channel
            if ap_static:
                vap_static[vap_tuple] += 1

            #   ClientMatch
            if ap_cm:
                vap_cm[vap_tuple] += 1
            if ap_cm_custom:
                vap_cm_custom[vap_tuple] += 1

        #global_vap_set[apn] = vap_tuple_set       # list of active VAPs on the AP




#############################################################################

#
#   Create Excel 3 (AP name vs VAP)
#

vap_list = {}       # key=vap, val=column number

# col = 6     # col1 = AP Name, col2 = AP model, col3 = Num of SSIDs, col4 = mixed, col5 = AP Group
# for apn in apn_sorted:
#     for vap in sorted(global_vap_set[apn.lower()]):
#         if vap in vap_list:
#             continue
#         vap_list[vap] = col
#         col += 1

#   Order VAP by # of APs
vap_sorted = sorted(global_vap_count.keys(), key=lambda x: global_vap_count[x], reverse=True)
col = 6
for vap in vap_sorted:
    vap_list[vap] = col
    col += 1

vap_header = [ f"{x[0]} ({x[2][0].upper()}) [{global_vap_count[x]}]" for x in vap_list.keys() ]
vap_fwd_mode = [ x[2] for x in vap_list.keys() ]
vap_cols = len(vap_header)

results = []

def row_add(func_name, vap_func):
    global vap_sorted, results
    row = [func_name]
    for vap in vap_sorted:
        if vap not in vap_func:
            row.append('-')
            continue
        n = vap_func[vap]
        if n == global_vap_count[vap]:
            row.append('〇')
        else:
            row.append(n)
    results.append(row)


#   MAC auth
row_add('MAC Authentication', vap_macauth)

#   802.1x reauth
row_add('802.1X Reauthentication', vap_reauth)

#   delete-keycache
row_add('Delete keycache', vap_del_keycache)

#   deny-inter-user-traffic
row_add('Deny Inter User Traffic', vap_deny_int_user)

#   Broadcast Filter
row_add('Broadcast Filter', vap_bcast_filter)

#   XML-API
row_add('XML API', vap_xmlapi)

#   Radius Accounting
row_add('Radius Accounting', vap_acct)

#   Interim Accounting
row_add('Radius Interim Accounting', vap_interim_acct)

#   User Derivation
row_add('User Derivation Rules', vap_user_deriv)

#   Stealth
row_add('Stealth SSID', vap_stealth)

#   Local Probe Threshold
row_add('802.11 Probe Threshold', vap_probe_thresh)

#   Auth Threshold
row_add('802.11 Auth Threshold', vap_auth_thresh)

#   Advertise AP name
row_add('Advertise AP name', vap_adv_apname)

#   VHT disabled
row_add('VHT(11ac) disabled', vap_vht_disable)

#   HT disabled
row_add('HT(11n) disabled', vap_ht_disable)

#   5GHz no channel bonding
row_add('5GHz No channel bonding', vap_no_bonding)

#   L2 Auth Fail Through
row_add('L2 Auth Fail Through', vap_l2_fail_thru)

#   Cellular Handoff Assist
row_add('Cellular Handoff Assist', vap_cha)

#   ClientMatch
row_add('ClientMatch', vap_cm)
row_add('CM custom parameter', vap_cm_custom)

#   Radio
row_add('2.4GHz disabled', vap_11g_off)
row_add('5GHz disabled', vap_11a_off)

#   Band
row_add('5GHz W52/W53/W56', vap_reg_all)
row_add('W52 only', vap_w52)
row_add('W53 only', vap_w53)
row_add('W56 only', vap_w56)
row_add('W52/W53 only', vap_w5253)
#row_add('W52/W56 only', vap_w5256)
row_add('W53/W56 only', vap_w5356)
row_add('5GHz custom', vap_reg_custom)

#   Static channel
row_add('Static channel assignment', vap_static)


######################################

df = pd.DataFrame(results, columns=['Functions'] + vap_header)

wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

f = Font(name='Consolas', size=9)
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.font = f

#   Header format
f = {
    'tunnel': PatternFill(fgColor='77CCFF', fill_type='solid'),
    'split-tunnel': PatternFill(fgColor='66EECC', fill_type='solid'),
    'bridge': PatternFill(fgColor='FFAAAA', fill_type='solid')
}
al = openpyxl.styles.Alignment(textRotation=90, horizontal='center')
for i, cell in enumerate(ws[1]):
    if i < 1:
        continue        # skip column A
    cell.alignment = al
    cell.fill = f[vap_fwd_mode[i-1]]

#   column widths
ws.column_dimensions['A'].width = 30

for i in range(vap_cols):
    ws.column_dimensions[get_column_letter(i+2)].width = 2

#   Center alignment
al = openpyxl.styles.Alignment(horizontal='center')
for r in ws.iter_rows(min_row=2, min_col=2):
    for cell in r:
        cell.alignment = al

#   freeze
ws.freeze_panes = "B2"

# filter
#ws.auto_filter.ref = "A:" + get_column_letter(ws.max_column)
#ws.auto_filter.ref = ws.dimensions

#
#   output to file
#
print(f"Writing to {xlsfile} ... ", end="")
wb.save(xlsfile)
print("done.")
