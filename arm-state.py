#
#   arm-state.py
#
#   show ap arm state をパースし、以下の条件の AP をカウント、coch-aps.xlsx に書き出す
#       - Neighbor AP (SNR が 10 以上で検出された Aruba AP)
#       - Coverage AP (SNR が 30 以上で検出された Aruba AP)
#       - Co-channel AP (SNR が 10 以上で検出され、チャネルが重複する Aruba AP)
#
#   複数コントローラが指定された場合、Neighbor AP 数が多い方のデータを残す(TBI)

import sys
import re
import argparse
import fileinput
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from collections import defaultdict
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from utils import isintf

COV_SNR = 30
xlsfile = "coch-aps.xlsx"


def fln():
    return fileinput.filename() + ":" + str(fileinput.filelineno())

apn2model = {}
apn2group = {}

grp2flr = {'Pasaraya 2nd Fl': 2,
           'Pasaraya 3rd Fl': 3,
           'Pasaraya 4th Fl': 4,
           'Pasaraya 5th Fl': 5,
           'Pasaraya 6th Fl': 6,
           'Pasaraya 7th Fl': 7,
           'Go-Jek Lantai 4': 4,
           'Go-Jek Lantai 5': 5,
           'Go-Jek Lantai 6': 6,
           'Go-Jek Lantai 7': 7,
           }
flrctr = defaultdict(lambda: 0)
flritf = defaultdict(lambda: 0)
grpctr = defaultdict(lambda: 0)
grpitf = defaultdict(lambda: 0)
allctr = 0
allitf = 0



def apn2floor(apn):
    pats = [r'(\d+|[GM])[fF]', r'B(\d+)-']
    # pats = [r'(G\dF\d)', r'(G\dM\d)', r'(G\dF.)', r'(..F\d)AP']     # for HKJC STRC
    for p in pats:
        m = re.search(p, apn)
        if m:
            return m.group(1)
    return 'n/a'

def parse_nbr_data(out, myapn, mych):
    global apn2model, apn2group
    global allctr, allitf
    global COV_SNR
    global enc

    if not re.search(args.pattern, myapn):
        return

    if myapn not in apn2group:
        log.warn(f"AP {myapn} not found in AP database")
        return

    cmd = out[0].strip()
    aos = AOSParser("".join(out), [cmd], encoding=enc)
    tbl = aos.get_table(cmd)
    if tbl is None or len(tbl) == 0:
        log.warn(fln() + f": No Neighbor Data found for AP {myapn}")
        print(f'{myapn},{apn2group[myapn]},{apn2model[myapn]},"{mych}",0,0')
        return [myapn, apn2group[myapn], apn2model[myapn], mych, 0, 0]

    nnbr = 0     # neighbor AP with SNR >= 10
    ncov = 0     # neighbor AP with SNR >= 30
    ncoch = 0    # neighbor AP with SNR >= 10 and overlapping channel

    for r in tbl[1:]:
        apn = r[0]
        if ':' in apn:
            # log.warn(f"skipping non-AP entry: {apn}")
            continue
        snr = int(r[2])
        if snr >= COV_SNR:
            ncov += 1
        if snr >= 10:
            nnbr += 1

            m = re.match(r'(\d+)[E+-]?/([\d\.]+)', r[5])
            if not m:
                m = re.match(r'(\d+)[E+-]?/([\d\.]+)', r[4])
            if m:
                ch = m.group(1)
                pwr = m.group(2)
            else:
                print(f"Can't parse Ch/EIRP: {fileinput.filelineno()}: {r[5]}")
                continue

            if isintf(args.band, ch, mych):
                ncoch += 1


    apg = apn2group[myapn]
    model = apn2model[myapn]

    # フロアごとに集計
    fl = apn2floor(myapn)
    flrctr[fl] += 1
    flritf[fl] += ncoch

    allctr += 1
    allitf += ncoch

    print(f'{myapn},{apg},{model},"{mych}",{nnbr},{ncov},{ncoch}')
    return [myapn, apg, model, mych, nnbr, ncov, ncoch]



#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Parse show ap tech and display neighbor APs")
    parser.add_argument('infile', help="Input file(s)", type=str, nargs='+')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    parser.add_argument('--pattern', '-p', help='Regex for AP name', type=str, default='.*')
    parser.add_argument('--band', '-b', help='Radio band', type=str, default='5')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)

    #
    #   parse ap database and get apname -> ap model mapping
    #
    enc = ""
    for enc in ('utf-8', 'mac-roman'):
        try:
            aos = AOSParser(args.infile, [AP_DATABASE_LONG_TABLE], merge=True, encoding=enc)
        except UnicodeDecodeError as e:
            continue
        break   # encode success
    else:
        print("unknown encoding, abort.")
        sys.exit(-1)

    ap_database_tbl = aos.get_table(AP_DATABASE_LONG_TABLE)
    if ap_database_tbl is None:
        print("show ap database long output not found.")
        sys.exit(-1)

    for r in ap_database_tbl[1:]:
        apn2model[r[0]] = r[2]
        apn2group[r[0]] = r[1]

    #
    #   parse show ap arm state
    #

    f = fileinput.input(args.infile, encoding=enc)
    for l in f:
        if l.startswith('show ap arm state'):
            break

    print("AP Name,Group,Type,Channel,Neighbor AP,Coverage AP,Co-ch AP")
    out = []
    apinfo = {}
    cont = False
    for l in f:
        if cont:
            if l.startswith('Legend: ') or l.startswith('AP:'):
                r = parse_nbr_data(out, apn, ch)        # myapn, apg, model, mych, nnbr, ncov, ncoch
                if r:
                    apn = r[0]
                    if apn not in apinfo or apinfo[apn][4] < r[4]:   # update if Neighbor AP count is higher
                        apinfo[apn] = r
                cont = False
            else:
                out.append(l)
                continue

        if not l.startswith('AP:'):
            continue

        r = re.match(r'AP:([\w-]+) MAC:[:\w]+ Band:([\w.]+) Channel:(\d+[SE+-]?)+', l)
        if r:
            apn = r.group(1)
            band = r.group(2)
            ch = r.group(3)
            if band[0] != args.band:
                continue
            cont = True
            out = ["show ap arm state\n"]
            continue

        continue


    #
    #   Average number of co-channel AP per floor
    #
    print("\nAverage number of co-channel APs per floor:")
    for fl in sorted(flrctr.keys()):
        avg = flritf[fl] / flrctr[fl]
        print(f"{fl:5}: {avg:6.2f} ({flrctr[fl]} APs)")

    print(f"Total avg co-ch APs: {allitf/allctr:.2f}")


    #
    #   Sort tbl
    #

    tbl = sorted(apinfo.values(), key=lambda x: x[0].lower())       # sort by AP name (case-insensitive)
    

    #
    #   Excel 書き出し
    #
    df = pd.DataFrame(tbl, columns=["AP Name", "Group", "Type", "Channel", "Neighbor AP", "Coverage AP", "Co-ch AP"])

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # set column width
    widths = [30, 40, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    # set header font and color
    f = Font(name='Calibri', bold=True, size=11, color="FFFFFF")
    p = PatternFill(fgColor="70AD47", fill_type="solid")
    for cell in ws['A1':'G1'][0]:
        cell.font = f
        cell.fill = p

    # set font, bgcolor
    f = Font(name='Consolas')
    p = PatternFill(fgColor="e2efda", fill_type="solid")
    rn = 1
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = f
            if rn&1:
                cell.fill = p
        rn += 1

    # centering for AP type
    for cell in ws['C'][1:]:
        cell.alignment = Alignment(horizontal='center')


    # set autofilter and freeze panes
    ws.auto_filter.ref = "A:G"
    ws.freeze_panes = "A2"


    #
    #   output to file
    #
    print(f"Writing to {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)
