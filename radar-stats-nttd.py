#!/usr/bin/python3 -u
#
#   show airmatch event radar パース for NTT-Data
#   AP 名、時刻で分類
#

import sys
import argparse
import re
import mylogger as log
from aos_parser import AOSParser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="parse tech-support and display client stats")
    parser.add_argument('files', type=str, nargs='*')
    parser.add_argument('--debug', help='debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)

    #
    #   parse datapath session table
    #
    #radar_cmd = "show airmatch event radar all-aps"
    radar_cmd = "show airmatch event .+"
    aos = AOSParser(args.files, radar_cmd, merge=True)

    #
    #   collect stats
    #
    radar_num = {}
    radar_tot = {}
    for r in aos.get_table(radar_cmd, "Event Type", "Timestamp", "APName", "Chan"):
        evt, ts, apn, ch = r
        if evt != "RADAR_DETECT":
            continue

        m = re.match("([a-zA-Z0-9]+)-", apn)
        if not m:
            log.err(f"Wrong AP name: {apn}")
            continue
        app = m.group(1)    # AP prefix

        m = re.match(r"\d\d\d\d-\d\d-\d\d_(\d\d):\d\d:\d\d", ts)
        if not m:
            log.err(f"Wrong timestamp: {ts}")
            continue        # timestamp format error
        hour = int(m.group(1))

        ch = int(ch)



        try:
            radar_num[app][hour] += 1
            radar_tot[app] += 1
        except KeyError:
            radar_num[app] = [0]*24
            radar_num[app][hour] = 1
            radar_tot[app] = 1

    #
    #   create Excel
    #
    wb = Workbook()
    ws = wb.active

    ws.append(["AP Prefix"] + list(range(0,24)) + ["Total"])

    for app in sorted(radar_tot.keys(), key=lambda x:radar_tot[x], reverse=True):
        #print(f"{apn:28}{radar_num[apn]} : ")
        ws.append([app] + radar_num[app] + [radar_tot[app]])

    #
    #   apply styles
    #
    ws.column_dimensions['A'].width = 21
    for col in "BCDEFGHIJKLMNOPQRSTUVWXY":
        ws.column_dimensions[col].width = 5

    f = Font(name='Calibri')
    for row in ws.iter_rows():
        for cell in row:
            cell.font = f

    f = Font(name='Calibri', bold=True)
    Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws[1]:
        cell.font = f
        cell.fill = Ses


    print(f"max_row={ws.max_row}")


    rs1 = str(ws.max_row)
    rs2 = str(ws.max_row+1)
    f = Font(name='Calibri')
    for col in "BCDEFGHIJKLMNOPQRSTUVWXY":
        sum = f"=SUM({col}2:{col}{rs1})"
        ws[col+rs2] = sum
        ws[col+rs2].font = f


    wb.save("radar-stats-nttd.xlsx")

    sys.exit(0)

