#!/usr/bin/python3 -u
#
#   show airmatch event radar パースし、radar-stats.xlsx に書き出す
#   --min <num> 回未満の AP は無視
#
#   以下のファイルも対応
#       - show ap arm history の結果を含むファイル
#       - Central の Events エクスポート csv

import sys
import argparse
import re
import mylogger as log
from aos_parser import AOSParser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
from utils import load_csv


MIN_THRESHOLD = 2

#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="parse radar events and export the summary to radar-stats.xlsx")
    parser.add_argument('files', type=str, nargs='*')
    parser.add_argument('--min', help='Minimum radar threshold', type=int, default=2)
    parser.add_argument('--fromdate', help='From date filter', type=str)
    parser.add_argument('--debug', help='debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)

    #
    #   parse radar event table
    #
    #radar_cmd = "show airmatch event radar all-aps"
    cmds = ["show airmatch event .+", "show ap arm history.*"]
    aos = AOSParser(args.files, cmds, merge=True)
    radar_tbl = aos.get_table(cmds[0], "APName", "Chan")
    armhist_tbl = aos.get_table(cmds[1], "filename", "Time of Change", "Old Channel", "New Channel", "Reason")


    if armhist_tbl:
        # arm history から radar event を抽出する
        radar_tbl = []
        radar_ts = []
        for r in armhist_tbl:
            fn, ts, chan, chan2, reason = r
            if reason == 'R':
                if args.fromdate and ts < args.fromdate:
                    continue
                m = re.search(r'([a-zA-Z0-9-]+)\.(log|txt)', fn)
                if m:
                    apn = m.group(1)
                else:
                    apn = fn.split(',')[0]
                pchan = re.sub(r'[SE+-]', '', chan)
                radar_tbl.append([apn, pchan])
                radar_ts.append((ts, apn, chan, chan2))
        for r in sorted(radar_ts):
            ts, apn, chan, chan2 = r
            print(f"{ts} {apn} {chan} -> {chan2}")



    elif radar_tbl is None:
        # fall back to csv (exported from Central events)
        tbl = load_csv(args.files[0])

        try:
            idx_name = tbl[0].index("Device Hostname")
        except ValueError:
            print("Error: 'Device Hostname' column not found in CSV file.")
            sys.exit(1)
        try:
            idx_desc = tbl[0].index("Description")
        except ValueError:
            print("Error: 'Description' column not found in CSV file.")
            sys.exit(1)

        radar_tbl = []
        for r in tbl[1:]:
            apn = r[idx_name]
            desc = r[idx_desc]
            m = re.search(r'channel (\d+)', desc)
            if m:
                ch = m.group(1)
                radar_tbl.append([apn, ch])



    #
    #   collect stats
    #
    radar_num = defaultdict(int)
    radar_ch = defaultdict(lambda: defaultdict(int))
    is144 = False
    for r in radar_tbl:
        apn, ch = r
        ch = int(ch)
        if ch == 144:
            is144 = True
        radar_num[apn] += 1
        radar_ch[apn][ch] += 1

    #
    #   create Excel
    #
    wb = Workbook()
    ws = wb.active

    dfs_ch = [52, 56, 60, 64, 100, 104, 108, 112, 116, 120, 124, 128, 132, 136, 140]
    if is144:
        dfs_ch.append(144)
    ws.append(["AP Name"] + dfs_ch + ["Total"])

    for apn in sorted(radar_num.keys(), key=lambda x:radar_num[x], reverse=True):
        #print(f"{apn:28}{radar_num[apn]} : ")
        if radar_num[apn] < args.min:
            break         # ignore APs with very few radar events
        nums = [radar_ch[apn][ch] if ch in radar_ch[apn] else 0 for ch in dfs_ch]
        ws.append([apn] + nums)

    #
    #   apply styles
    #

    #  set column width
    ws.column_dimensions['A'].width = 21
    if is144:
        ch_cols = "BCDEFGHIJKLMNOPQ"
    else:
        ch_cols = "BCDEFGHIJKLMNOP"
    for col in ch_cols:
        ws.column_dimensions[col].width = 5

    #   set font
    f = Font(name='Calibri')
    for row in ws.iter_rows():
        for cell in row:
            cell.font = f

    f = Font(name='Calibri', bold=True)
    Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws[1]:
        cell.font = f
        cell.fill = Ses

    #   apply conditional formatting
    f1 = PatternFill(fgColor="FFE5E8", fill_type="solid")
    f2 = PatternFill(fgColor="FFC7CE", fill_type="solid")
    f3 = PatternFill(fgColor="FDB58D", fill_type="solid")
    f4 = PatternFill(fgColor="FF6600", fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=len(dfs_ch)+1):
        for cell in row:
            if cell.value >= 20:
                cell.fill = f4
            elif cell.value >= 10:
                cell.fill = f3
            elif cell.value >= 5:
                cell.fill = f2
            elif cell.value > 0:
                cell.fill = f1

    #   add total column
    f = Font(name='Calibri', bold=True)
    for row in range(2,ws.max_row+1):
        rs = str(row)
        sum = f"=SUM({ch_cols[0]}{rs}:{ch_cols[-1]}{rs})"
        sum_col = chr(ord(ch_cols[-1])+1)
        ws[sum_col + rs] = sum
        # bold font
        ws[sum_col + rs].font = f


    rs1 = str(ws.max_row)
    rs2 = str(ws.max_row+1)
    f = Font(name='Calibri', bold=True)
    for col in ch_cols:
        sum = f"=SUM({col}2:{col}{rs1})"
        ws[col+rs2] = sum
        ws[col+rs2].font = f


    wb.save("radar-stats.xlsx")
    print("Radar stats saved to radar-stats.xlsx")

    sys.exit(0)

