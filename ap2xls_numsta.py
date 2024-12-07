#/usr/bin/python3 -u
#
#   ap2xls_numsta.py
#
#   Join 'show ap database' and 'show ap active' tables and write the result to an MS Excel file
#
#   2024/10/24 - AP-655 (tri-radio) 対応

import sys
import re
import argparse
import pandas as pd
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

# APpat = "^APGTS"
# APpat = "^APG7"
# APpat = "^APUMEDA"
# APpat = "^APHIBFS"
# APpat = "^MICL5"
# APpat = "^E013-A02[3456]"
# APpat = r'SG-WA-F02|SG-WC-F02'
# APpat = r'^wse'
APpat = r'^(hvnap[0-9b]+|HVNAP[0-9b]+)'

def uniq(tbl):
    ret = []
    k = set()
    for r in tbl:
        if r[0] in k: continue
        k.add(r[0])
        ret.append(r)
    return ret

def toi(s):
    try:
        return int(s)
    except ValueError:
        return 0


#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Join 'show ap database' and 'show ap active' tables and write the result to an MS Excel file")
    parser.add_argument('infile', help="Input file(s) containing 'show ap database long' and 'show ap active' output", type=str, nargs='+')
    parser.add_argument('outfile', help='Output Excel file', type=str, nargs='?', default='')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)

    if args.outfile == '':
        xlsfile = 'ap-table.xlsx'
    else:
        xlsfile = args.outfile

    #
    #   parse AP tables
    #
    print("Parsing files ... ", end="")
    aos = AOSParser(args.infile, [AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE], merge=True)
    ap_database_tbl = aos.get_table(AP_DATABASE_LONG_TABLE)
    ap_active_tbl   = aos.get_table(AP_ACTIVE_TABLE)
    if ap_database_tbl is None:
        print("show ap database long output not found.")
        sys.exit(-1)
    if ap_active_tbl is None:
        print("show ap active output not found.")
        sys.exit(-1)
    print("done.")

    ap_database_tbl = uniq(ap_database_tbl)
    print(f"{len(ap_database_tbl)-1} unique APs found in ap database.")
    ap_active_tbl = uniq(ap_active_tbl)
    print(f"{len(ap_active_tbl)-1} unique APs found in active ap table.")

    #
    #   filter AP database table
    #
    ap_db_tbl = [ap_database_tbl[0]]
    for r in ap_database_tbl[1:]:
        if 'APpat' in globals() and not re.search(APpat, r[0]):
            continue
        ap_db_tbl.append(r)

    #
    #   AP model tally
    #
    apmodelctr = defaultdict(lambda: 0)
    for r in ap_db_tbl[1:]:
        model = r[2]
        apmodelctr[model] += 1
    print("AP models")
    for model in sorted(apmodelctr.keys()):
        print(f"{model}: {apmodelctr[model]}")
    print()


    #
    #   Process columns in active AP table
    #
    ap_act_tbl = [
        ["Name", "5G PHY", "5G Ch", "5G EIRP", "5G STA", "2.4G PHY", "2.4G Ch", "2.4G EIRP", "2.4G STA",
         "6G PHY", "6G Ch", "6G EIRP", "6G STA"]]
    idx_r0 = ap_active_tbl[0].index("Radio 0 Band Ch/EIRP/MaxEIRP/Clients")
    idx_r1 = ap_active_tbl[0].index("Radio 1 Band Ch/EIRP/MaxEIRP/Clients")
    try:
        idx_r2 = ap_active_tbl[0].index("Radio 2 Band Ch/EIRP/MaxEIRP/Clients")
    except ValueError:
        idx_r2 = -1

    num_ch = defaultdict(lambda: 0)
    sta_ch = defaultdict(lambda: 0)
    usersperfloor = defaultdict(lambda: 0)
    apsperfloor = defaultdict(lambda: 0)

    for row in ap_active_tbl[1:]:
        apn = row[0]
        if 'APpat' in globals() and not re.search(APpat, apn):
            continue

        r0 = row[idx_r0]
        r = re.match(r"(.+):([\dSE+\-]+)/([\d\.]+)/[\d\.]+/(\d+)$", r0)      # AP:5GHz-HE:124/14.0/30.0/8
        if r:
            r0_phy  = r.group(1)
            r0_ch   = r.group(2)
            r0_eirp = float(r.group(3))
            r0_sta  = int(r.group(4))
            num_ch['5G:'+r0_ch] += 1
            sta_ch['5G:'+r0_ch] += r0_sta
        else:
            r0_phy = r0_ch = r0_eirp = r0_sta = ""

        r1 = row[idx_r1]
        r = re.match(r"(.+):([\dSE+\-]+)/([\d\.-]+)/[\d\.-]+/(\d+)$", r1)      # AP:2.4GHz-HE:11/8.0/27.7/0
        if r:
            r1_phy  = r.group(1)
            r1_ch   = r.group(2)
            r1_eirp = float(r.group(3))
            r1_sta  = int(r.group(4))
            num_ch['2.4G:'+r1_ch] += 1
            sta_ch['2.4G:'+r1_ch] += r1_sta
        else:
            r1_phy = r1_ch = r1_eirp = r1_sta = ""

        if idx_r2 == -1:
            r2_phy = r2_ch = r2_eirp = r2_sta = ""
        else:
            r2 = row[idx_r2]
            r = re.match(r"(.+):([\dSE+\-]+)/([\d\.]+)/[\d\.]+/(\d+)$", r2)      # AP:6GHz-HE:85E/15.0/24.0/7
            if r:
                r2_phy  = r.group(1)
                r2_ch   = r.group(2)
                r2_eirp = float(r.group(3))
                r2_sta  = int(r.group(4))
                num_ch['6G:'+r2_ch] += 1
                sta_ch['6G:'+r2_ch] += r2_sta
            else:
                r2_phy = r2_ch = r2_eirp = r2_sta = ""

        ap_act_tbl.append([apn, r0_phy, r0_ch, r0_eirp, r0_sta, r1_phy, r1_ch, r1_eirp, r1_sta, r2_phy, r2_ch, r2_eirp, r2_sta])

        #
        #  フロア毎のユーザ数集計
        #
        # m = re.match(r'(\w+\d\d)', row[1])
        m = re.search(r'hvnap([0-9b]+)fap', apn, re.IGNORECASE)
        if m:
            fl = m.group(1)
        else:
            fl = 'n/a'
            #print(f"Uknown floor: AP {row[0]} Group {row[1]}")

        apsperfloor[fl] += 1
        usersperfloor[fl] += toi(r0_sta) + toi(r1_sta) + toi(r2_sta)


    print("Channel distribution")
    print("channel: radios (clients)")
    for ch in sorted(num_ch.keys()):
        print(f"{ch}: {num_ch[ch]} ({sta_ch[ch]})")

    print("\nUsers/APs per floor")
    for fl in sorted(usersperfloor.keys()):
        print(f"{fl}: {usersperfloor[fl]} users/{apsperfloor[fl]} APs")


    #
    #   table join & sort
    #
    df_ap_database = pd.DataFrame(ap_db_tbl[1:], columns=ap_db_tbl[0])
    df_ap_act      = pd.DataFrame(ap_act_tbl[1:], columns=ap_act_tbl[0])

    df1 = df_ap_database.merge(df_ap_act, how="left", left_on="Name", right_on="Name")
    df2 = df1[[
        'Name', 'Group', 'AP Type', 'IP Address', 'Switch IP',
        '5G PHY', '5G Ch', '5G EIRP', '5G STA',
        '2.4G PHY', '2.4G Ch', '2.4G EIRP', '2.4G STA',
        '6G PHY', '6G Ch', '6G EIRP', '6G STA',
        'Serial #', 'Wired MAC Address',
    ]]
    #df = df2.sort_values(['Group', 'Name'])
    df = df2.sort_values(['Name'])

    #
    #   Create Excel
    #
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    f = Font(name='Consolas')
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.font = f

    widths = [25, 30, 10, 20, 20,   15, 10, 10, 10,   15, 10, 10, 10,   15, 10, 10, 10,  13, 25]
    for i,w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    f = Font(name='Arial', bold=True, size=9)
    Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws['A1':'S1'][0]:
        cell.fill = Ses
        cell.font = f

    ws.auto_filter.ref = "A:S"
    ws.freeze_panes = "A2"


    #
    #   output to file
    #
    print(f"Writing to {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)
