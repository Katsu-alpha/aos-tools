#/usr/bin/python3 -u
#
#   AP database を Excel に保存
#

import sys
import argparse
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="convert 'show ap database long' output to MS Excel")
    parser.add_argument('infile', help='input file containing show ap database long output', type=str, nargs=1)
    parser.add_argument('outfile', help='output Excel file', type=str, nargs='?', default='')
    parser.add_argument('--debug', help='enable debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)

    if args.outfile == '':
        xlsfile = 'ap-database.xlsx'
    else:
        xlsfile = args.outfile

    #
    #   parse AP tables
    #
    print("Parsing AP Database ... ", end="")
    aos = AOSParser(args.infile, AP_DATABASE_LONG_TABLE)
    hdrs = ["Name", "Group", "AP Type", "IP Address", "Switch IP", "Wired MAC Address", "Serial #"]
    tbl = aos.get_table(AP_DATABASE_LONG_TABLE, *hdrs)
    print("done.")

    #
    #   Create Excel
    #
    wb = Workbook()
    ws = wb.active
    ws.append(hdrs)
    for r in tbl:
        ws.append(r)

    f = Font(name="Consolas", size=10)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = f

    widths = [40, 25, 8, 15, 15, 18, 15]
    for i,w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    f = Font(name="Consolas", size=10, bold=True)
    s = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws['A1':'G1'][0]:
        cell.fill = s
        cell.font = f

    ws.auto_filter.ref = "A:G"
    ws.freeze_panes = "A2"

    #
    #   output to file
    #
    print(f"Writing {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)

