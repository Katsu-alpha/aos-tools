#
#   ap2xls-iap.py
#
#   IAP/AOS10 AP の show tech-support をパースし、以下の情報を Excel に書き出し
#       - AP Name, Model, Version, Uptime
#       - 各Radioについて、PHY, Channel, EIRP, Clients, Util(%), OBSS(%), Intf(%), Noise(dBm)
#

import re
import argparse
import sys
import mylogger as log
import pandas as pd
from aos_parser import AOSParser
from collections import defaultdict
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


RADIO_BAND = ['5', '2', '6']   # Radio 0/wifi0=5GHz, Radio 1/wifi1=2.4GHz, Radio 2/wifi2=6GHz

def avg(it):
    s = 0
    n = 0
    for v in it:
        s += v
        n += 1
    return s//n if n > 0 else 0

parser = argparse.ArgumentParser(
    description="Parse IAP/AOS10 AP tech-support and generate Excel file")
parser.add_argument('infiles', help="Input file(s) containing 'show tech-support' output", type=str, nargs='+')
parser.add_argument('--debug', help='Enable debug log', action='store_true')
args = parser.parse_args()

xlsfile = 'aplist.xlsx'

if args.debug:
    log.setloglevel(log.LOG_DEBUG)
else:
    log.setloglevel(log.LOG_INFO)

cmds = ['show ap association', 'show ap bss-table']


aps = set()

numsta = defaultdict(lambda: defaultdict(int))
htmode = defaultdict(lambda: defaultdict(str))

channel = defaultdict(lambda: defaultdict(str))
eirp = defaultdict(lambda: defaultdict(str))

uptime = defaultdict(str)
ibss = defaultdict(lambda: defaultdict(lambda: None))
obss = defaultdict(lambda: defaultdict(lambda: None))
intf = defaultdict(lambda: defaultdict(lambda: None))

nf = defaultdict(lambda: defaultdict(lambda: None))
busy = defaultdict(lambda: defaultdict(lambda: None))

model = defaultdict(str)
osver = defaultdict(str)

for fn in args.infiles:
    print(f"Processing {fn}")

    aos = AOSParser(fn, cmds)

    assoc_tbl = aos.get_table(cmds[0], 'assoc', 'phy')
    if assoc_tbl is None:
        assoc_tbl = []
    bss_tbl = aos.get_table(cmds[1], 'band/ht-mode/bandwidth', 'ch/EIRP/max-EIRP', 'ap name')
    if bss_tbl is None:
        print("show ap bss-table output not found.")
        continue

    #   parse bss-table and get AP Name, HT mode, Ch, EIRP
    maxbw = defaultdict(lambda: defaultdict(int))
    for r in bss_tbl:
        band_cbw, ch_eirp, apn = r
        m = re.match(r'([\d.]+)GHz/(\w+)/(\d+)MHz', band_cbw)
        if not m:
            print(f"Invalid band/ht-mode/bandwidth: {band_cbw}")
            sys.exit(1)
        cbw = int(m.group(3))
        band = m.group(1)[0]   # "5GHz" -> '5'
        if cbw <= maxbw[band][apn]:
            continue
        maxbw[band][apn] = cbw
        htmode[band][apn] = band + 'GHz-' +m.group(2)
        channel[band][apn], eirp[band][apn] = ch_eirp.split('/')[:2]

    # avoid double counting
    if apn in aps:
        print(f"Duplicate AP Name found: {apn}")
        sys.exit(1)
    aps.add(apn)


    # parse assoc table and get number of clients per radio
    for r in assoc_tbl:
        assoc, phy = r
        if assoc != 'y':
            continue
        if phy[0] in ('2', '5', '6'):
            numsta[phy[0]][apn] += 1


    pat = re.compile(r"(^AP Uptime$|^CCA stats history:|show ap debug radio-stats|show version|^wifi[012]     phy_stats:)")
    #
    #   parse file line by line
    #
    f = open(fn, 'r', encoding='macroman')
    while (l := f.readline()):
        if not pat.search(l):
            continue
        if l.startswith('end of show '):
            continue

        if l.startswith("AP Uptime"):
            _ = f.readline()   # skip next line
            uptime[apn] = ' '.join(f.readline().strip().split()[:2])
            continue

        # Get obss(%)
        #   - show ap arm rf-summary
        #   - last 30 samples, 1 sample per 1 second, latest first
        #    CCA stats history:wifi0
        #   --- example output ---
        #    Phy-Type:5GHz
        #    ch:      52   52   52   52   52   52   52   52   52   52   52   52   52   52
        #    ibss:     1    2    2    2    2    3    4    6    6    3    1    2    3    4
        #    obss:     3    3    3    3    3    3    4    3    3    4    4    4    4    4
        #    intf:     0    0    0    0    0    0    0    0    0    0    0    0    0    0
        #   ----------------------
        #    
        #   以下のバージョンには存在しない
        #   10.4.1.5 の AP635
        if l.startswith("CCA stats history:"):
            if "wifi0" in l:
                radio = RADIO_BAND[0]
            elif "wifi1" in l:
                radio = RADIO_BAND[1]
            elif "wifi2" in l:
                radio = RADIO_BAND[2]
            else:
                continue
            while True:
                l = f.readline().strip()
                if not l:
                    break
                if l.startswith("ibss: "):
                    ibss[radio][apn] = avg(map(int, l[5:].strip().split()))     # max -> avg
                    continue
                if l.startswith("obss: "):
                    obss[radio][apn] = avg(map(int, l[5:].strip().split()))     # max -> avg
                    continue
                if l.startswith("intf: "):
                    intf[radio][apn] = avg(map(int, l[5:].strip().split()))     # max -> avg
                    break
            continue

        # Get obss(%), intf(%)
        #   - show ap debug radio-info
        #   - last 10 samples, 1 sample per 1 seconds, latest last
        #   --- example output ---
        #   wifi0     phy_stats:0
        #   
        #   Tx Time (%)             1        1        1        1        1        1        4        1        1        1
        #   Rx Time (%)             3        3        3        4        3        3        3        3        3        3
        #   Rx Time To Me (%)       0        0        0        0        0        0        0        0        0        0
        #   CCA Busy (%)            5        4        5        6        5        4        8        5        5        5
        #   	                   ------------------------------------------------------------------------------------------
        #   Interference (%)        1        0        1        1        1        0        1        1        1        1
        #   ----------------------
        #
        #   BRCM AP では出力されない
        if l.startswith("wifi"):
            radio = RADIO_BAND[int(l[4])]
            for _ in range(7):
                l = f.readline()
                if l.startswith("Rx Time (%) "):
                    rx_time = avg(map(int, l[12:].strip().split()))
                    continue
                if l.startswith("Rx Time To Me (%) "):
                    rx_time_to_me = avg(map(int, l[18:].strip().split()))
                    continue
                if l.startswith("Interference (%) "):
                    _intf = avg(map(int, l[17:].strip().split()))
                    break
            if apn not in intf[radio]:
                intf[radio][apn] = _intf
            if apn not in obss[radio]:
                obss[radio][apn] = rx_time - rx_time_to_me
            continue

        if "show ap debug radio-stats" in l:
            if "radio-stats 0" in l:
                radio = RADIO_BAND[0]
            elif "radio-stats 1" in l:
                radio = RADIO_BAND[1]
            elif "radio-stats 2" in l:
                radio = RADIO_BAND[2]
            else:
                continue

            while True:
                l = f.readline()
                if not l or l.startswith("Command Failed") or l.startswith("end of show ap debug radio-stats"):
                    break

                if l.startswith("Current Noise Floor"):
                    nf[radio][apn] = -int(l.split()[-1])
                    continue
                if l.startswith("Channel Busy 64s"):
                    busy[radio][apn] = int(l.split()[-1])
                    break
            continue

        if "show version" in l:
            for _ in range(3):
                l = f.readline()
                if m:=re.search(r"MODEL: ([\w\d]+)", l):
                    model[apn] = m.group(1).strip()
                    if m:=re.search(r"Version ([\d.]+)", l):
                        osver[apn] = m.group(1).strip()
                    break
            continue


#
#   Create a dataframe
#

tbl = []
for apn in sorted(aps):
    row = [apn, model[apn], osver[apn], uptime[apn],
            htmode['2'][apn], channel['2'][apn], eirp['2'][apn], numsta['2'][apn], busy['2'][apn], obss['2'][apn], intf['2'][apn], nf['2'][apn],
            htmode['5'][apn], channel['5'][apn], eirp['5'][apn], numsta['5'][apn], busy['5'][apn], obss['5'][apn], intf['5'][apn], nf['5'][apn],
            htmode['6'][apn], channel['6'][apn], eirp['6'][apn], numsta['6'][apn], busy['6'][apn], obss['6'][apn], intf['6'][apn], nf['6'][apn],
            ]
    tbl.append(row)

df = pd.DataFrame(tbl, columns=['AP Name', 'Model', 'AOS Version', 'Uptime',
                                'Mode', 'Channel', 'EIRP(dBm)', 'Clients', 'Util(%)', 'OBSS(%)', 'Intf(%)', 'Noise(dBm)',
                                'Mode', 'Channel', 'EIRP(dBm)', 'Clients', 'Util(%)', 'OBSS(%)', 'Intf(%)', 'Noise(dBm)',
                                'Mode', 'Channel', 'EIRP(dBm)', 'Clients', 'Util(%)', 'OBSS(%)', 'Intf(%)', 'Noise(dBm)',
                                ])


#
#   Write to Excel
#
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

f = Font(name='Consolas')
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.font = f

widths = [25, 10, 15, 10,   10, 10, 10,    10, 10, 10, 10,   10, 10, 10,   10, 10, 10, 10,   10, 10, 10,   10, 10, 10, 10]
for i,w in enumerate(widths):
    ws.column_dimensions[chr(65+i)].width = w

f = Font(name='Arial', bold=True, size=9)
Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
for cell in ws['A1':'AB1'][0]:
    cell.fill = Ses
    cell.font = f

ws.auto_filter.ref = "A:AB"
ws.freeze_panes = "A2"


#
#   output to file
#
print(f"Writing to {xlsfile} ... ", end="")
wb.save(xlsfile)
print("done.")

sys.exit(0)
