#
#   show run summrize
#
#   APグループ数 (使用・未使用)
#   Virtual AP 数 (使用・未使用)
#   AP model 一覧
#

import re
import glob
import sys
import argparse
import mylogger as log
from aos_parser import AOSParser
from colorama import Fore, Style
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def print_debug(msg):
    print(Fore.GREEN + msg + Style.RESET_ALL)

def print_info(msg):
    print(Fore.CYAN + msg + Style.RESET_ALL)

def print_warn(msg):
    print(Fore.YELLOW + msg + Style.RESET_ALL)

def print_err(msg):
    print(Fore.RED + msg + Style.RESET_ALL)


class SSID:
    def __init__(self, name):
        self.name = name
        self.ssid = ''
        self.opmode = 'open'
        self.stealth = False
        self.disabled = False

class VAP:
    def __init__(self, name):
        self.name = name
        self.ssid_prof = None
        self.aaa_prof = None
        self.fwd_mode = 'tunnel'
        self.vlan = 1
        self.disabled = False

class APGroup:
    def __init__(self, name):
        self.name = name
        self.vaps = []
        self.dot11a = None
        self.dot11g = None

class APName:
    def __init__(self, name):
        self.name = name
        self.vaps = []
        self.exclude_vaps = []
        self.dot11a = None
        self.dot11g = None
        self.reg = None
        self.misc = []

class ARM:
    def __init__(self, name):
        self.name = name
        self.disabled = False

class DOT11A:
    def __init__(self, name):
        self.name = name
        self.arm_prof = None
        self.radio = True

class DOT11G:
    def __init__(self, name):
        self.name = name
        self.arm_prof = None
        self.radio = True


class AOSConfig:
    def __init__(self, name):
        self.name = name        # ホスト名
        self.num_user_roles = 0 # user-role 数
        self.num_acls = 0       # session ACL 数
        self.ssid_profs = {}    # key=プロファイル名, val=SSID obj
        self.vap_profs = {}     # key=プロファイル名, val=VAP obj
        self.ap_groups = {}     # key=AP グループ名, val=APGroup obj
        self.ap_names = {}      # key=AP 名, val=APName obj
        self.arm_profs = {}
        self.dot11a_profs = {}
        self.dot11g_profs = {}


def parse_config(name, config):
    C = AOSConfig(name)

    dot11a_def = DOT11A('default')
    C.dot11a_profs['default'] = dot11a_def
    dot11g_def = DOT11G('default')
    C.dot11g_profs['default'] = dot11g_def

    in_cont = False

    for l in config:
        if not in_cont:

            if l.startswith('ip access-list session '):
                C.num_acls += 1  # ip access-list 数カウント
                continue

            if l.startswith('user-role '):
                C.num_user_roles += 1  # user-role 数カウント
                continue

            if l.startswith('ap system-profile '):
                in_cont = 'apsys'
                name = l[18:].replace('"', '')
                continue

            if l.startswith('wlan ssid-profile '):
                in_cont = 'ssid'
                name = l[18:].replace('"', '')
                ssid = SSID(name)
                C.ssid_profs[name] = ssid
                continue

            if l.startswith('wlan virtual-ap '):
                in_cont = 'vap'
                name = l[16:].replace('"', '')
                vap = VAP(name)
                C.vap_profs[name] = vap
                continue

            if l.startswith('ap-group '):
                in_cont = 'apg'
                name = l[9:].replace('"', '').lower()
                apg = APGroup(name)
                C.ap_groups[name] = apg
                continue

            if l.startswith('ap-name '):
                in_cont = 'apn'
                name = l[8:].replace('"', '').lower()
                apn = APName(name)
                continue

            if l.startswith('rf arm-profile '):
                in_cont = 'arm'
                name = l[15:].replace('"', '')
                arm = ARM(name)
                C.arm_profs[name] = arm
                continue

            if l.startswith('rf dot11a-radio-profile '):
                in_cont = 'dot11a'
                name = l[24:].replace('"', '')
                dot11a = DOT11A(name)
                C.dot11a_profs[name] = dot11a
                continue

            if l.startswith('rf dot11g-radio-profile '):
                in_cont = 'dot11g'
                name = l[24:].replace('"', '')
                dot11g = DOT11G(name)
                C.dot11g_profs[name] = dot11g
                continue

            continue

        #
        #   inside a context
        #

        l = l.strip()
        if l == '!':
            in_cont = False
            continue

        if in_cont == 'apsys':
            r = re.match(r'dns-domain |lms-ip |bkup-lms-ip |bootstrap-threshold |number_ipsec_retries |secondary-master |shell-passwd |bkup-passwords |lms-preemption|ap-console-password |heartbeat-interval |session-acl ', l)
            if not r:
                print_warn(f'ap system-profile {name}: {l}')   # 特殊な system-profile 設定
                continue
            continue

        if in_cont == 'ssid':
            if l.startswith('essid '):
                ssid.ssid = l[6:].replace('"', '')
                continue
            if l.startswith('opmode '):
                ssid.opmode = l[7:]
                continue
            if l.startswith('hide-ssid'):
                ssid.stealth = True
                continue
            if l.startswith('no ssid-enable'):
                ssid.disabled = True
                continue
            continue

        if in_cont == 'vap':
            if l.startswith('aaa-profile '):
                vap.aaa_prof = l[12:].replace('"', '')
                continue
            if l.startswith('ssid-profile '):
                n = l[13:].replace('"', '')
                vap.ssid_prof = C.ssid_profs[n]
                if vap.ssid_prof.disabled:
                    vap.disabled = True
                continue
            if l.startswith('vlan '):
                vap.vlan = int(l[5:])
                continue
            if l.startswith('forward-mode '):
                vap.fwd_mode = l[13:]
                continue
            r = re.match(r'allowed-band |broadcast-filter |deny-inter-user-traffic|cellular-handoff-assist', l)
            if not r:
                print_warn(f'virtual-ap {name}: {l}')
            continue

        if in_cont == 'apg':
            if l.startswith('virtual-ap '):
                n = l[11:].replace('"', '')
                apg.vaps.append(C.vap_profs[n])
                C.ap_groups[name] = apg
                continue
            if l.startswith('dot11a-radio-profile '):
                n = l[21:].replace('"', '')
                apg.dot11a = C.dot11a_profs[n]
                continue
            if l.startswith('dot11g-radio-profile '):
                n = l[21:].replace('"', '')
                apg.dot11g = C.dot11g_profs[n]
                continue
            continue

        if in_cont == 'apn':
            if l.startswith('virtual-ap '):
                n = l[11:].replace('"', '')
                apn.vaps.append(C.vap_profs[n])
                C.ap_names[name] = apn
                continue
            if l.startswith('exclude-virtual-ap '):
                n = l[19:].replace('"', '')
                apn.exclude_vaps.append(n)
                C.ap_names[name] = apn
                continue
            if l.startswith('dot11a-radio-profile '):
                n = l[21:].replace('"', '')
                apn.dot11a = C.dot11a_profs[n]
                continue
            if l.startswith('dot11g-radio-profile '):
                n = l[21:].replace('"', '')
                apn.dot11g = C.dot11g_profs[n]
                continue
            if l.startswith('regulatory-domain-profile '):
                n = l[26:].replace('"', '')
                apn.reg = n
                continue
            #print_warn(f'ap-name {name}: {l}')
            apn.misc.append(l)
            continue

        if in_cont == 'arm':
            if l.startswith('assignment disable'):
                arm.disabled = True
                continue
            continue

        if in_cont == 'dot11a':
            if l.startswith('no radio-enable'):
                dot11a.radio = False
                continue
            if l.startswith('arm-profile '):
                n = l[12:].replace('"', '')
                try:
                    dot11a.arm_prof = C.arm_profs[n]
                except KeyError:
                    print_err(f"In dot11a radio profile '{name}'")
                    print_err(f"ARM profile '{n}' referenced but not defined.")
                continue
            continue

        if in_cont == 'dot11g':
            if l.startswith('no radio-enable'):
                dot11g.radio = False
                continue
            if l.startswith('arm-profile '):
                n = l[12:].replace('"', '')
                try:
                    dot11g.arm_prof = C.arm_profs[n]
                except KeyError:
                    print_err(f"In dot11g radio profile '{name}'")
                    print_err(f"ARM profile '{n}' referenced but not defined.")
                continue
            continue

    return C    # return an AOSConfig object

#-------------------------------------------------------------------------

#
#   start
#

parser = argparse.ArgumentParser(
    description="parse show run/show ap database and summarize")
parser.add_argument('files', help="show run/show ap database long outputs", type=str, nargs='*')
parser.add_argument('--debug', help='debug log', action='store_true')
args = parser.parse_args()

if args.debug:
    log.setloglevel(log.LOG_DEBUG)
else:
    log.setloglevel(log.LOG_INFO)

AP_DATABASE_LONG = "show ap database long"

if '*' in args.files[0]:
    args.files = glob.glob(args.files[0])

xlsfile1 = 'summary-MC.xlsx'    # Controller summary
xlsfile2 = 'summary-SSID.xlsx'  # SSID summary
xlsfile3 = 'summary-ALL.xlsx'   # All AP/SSID detail

#
#   process each file
#
coex_vaps_sets = {}         # list of co-existing VAPs. key=VAP tuple(ssid, opmode, fwd_mode), val=set()
coex_vaps_count = {}        # co-ex VAPs count. key=VAP tuple, val=dict(key=co-ex vap, elem=count)
controller_sets = {}        # list of controllers hosting the VAP. key=(ssid, opmode, fwd_mode), element=set()
global_vap_count = {}       # key=VAP tuple
global_vap_set = {}         # key=AP Name(lower), val=VAP tuple
global_apn_set = set()      # names of Active APs
global_apn_apg_dic = {}     # AP name -> AP group mapping
global_mixed_ap = set()
global_ap_model = {}

results = []                # table for excel output
for fn in args.files:
    print(f"Processing file: {fn}")

    # Parse 'show ap database long' table
    aos = AOSParser(fn, cmds=[AP_DATABASE_LONG])

    ap_db_tbl = aos.get_table(AP_DATABASE_LONG)
    if ap_db_tbl is None:
        print_info(f"No active AP found in {fn}")
        continue

    ap_db_tbl = ap_db_tbl[1:]   # discard header

    #
    #   Count active AP groups/AP types, create ap-name/group sets
    #
    apg_set = set()
    apg_count = {}
    aptype_count = {}
    for r in ap_db_tbl:
        apn = r[0]
        global_apn_set.add(apn)
        grp = r[1]
        global_apn_apg_dic[apn] = grp

        grp_l = grp.lower()
        apg_set.add(grp_l)
        if grp_l not in apg_count:
            apg_count[grp_l] = 0
        apg_count[grp_l] += 1

        ap_type = r[2]
        if ap_type not in aptype_count:
            aptype_count[ap_type] = 0
        aptype_count[ap_type] += 1
        global_ap_model[apn] = ap_type

    #
    #   Parse running-config, get controller model, version
    #
    in_showrun = False
    config = []
    with open(fn) as f:
        for line in f:
            if in_showrun:
                if re.match(r'\([\w\-_]+\) #', line):   # AOS6 prompt
                    break
                if re.match(r'\([\w\-_]+\) \*?\[[\w/]+\] #', line):   # AOS8 prompt
                    break
                config.append(line.rstrip())
                continue
            if re.search("show running-config", line):
                in_showrun = True
                continue
            if re.search(r'show configuration committed /mm/mynode', line):
                in_showrun = True
                continue
            r = re.match(r'ArubaOS \(MODEL: (\w+)\), Version ([\d\.]+)', line)
            if r:
                model = r.group(1)
                version = r.group(2)
                continue

    controller = fn[:-4]
    c = parse_config(controller, config)

    #   flag used
    # for name, apg in c.ap_groups.items():
    #     if name in apg_used:
    #         apg.used = True
    #         for vap in apg.vaps:
    #             vap.used = True
    #             vap.ssid_prof.used = True
    #
    # for name, apn in c.ap_names.items():
    #     if name in apn_used:
    #         apn.used = True
    #         for vap in apn.vaps:
    #             vap.used = True
    #             vap.ssid_prof.used = True


    #   check static assignment, count used SSIDs
    num_static = 0
    vap_count = {}      # key=(ssid, opmode, fwd-mode)
    mixed_ap_count = 0
    for r in ap_db_tbl:
        apg = r[1].lower()
        apn = r[0].lower()

        dot11a = c.ap_groups[apg].dot11a
        dot11g = c.ap_groups[apg].dot11g
        dot11a_static = False
        dot11g_static = False

        if dot11a and dot11a.arm_prof and dot11a.arm_prof.disabled:
            dot11a_static = True
        if dot11g and dot11g.arm_prof and dot11g.arm_prof.disabled:
            dot11g_static = True

        both_radio_disabled = False
        if apn in c.ap_names:
            apn_c = c.ap_names[apn]
            dot11a = apn_c.dot11a
            dot11g = apn_c.dot11g
            if dot11a and dot11a.arm_prof and dot11a.arm_prof.disabled:
                dot11a_static = True
            if dot11g and dot11g.arm_prof and dot11g.arm_prof.disabled:
                dot11g_static = True
            if dot11a and dot11a.radio is False and dot11g and dot11g.radio is False:
                print_err(f"Both radios are disabled on {r[0]}!")
                both_radio_disabled = True

        if dot11a_static or dot11g_static:
            num_static += 1

        if both_radio_disabled:
            continue

        #   create VAP set for the AP
        vap_set = set()
        for vap in c.ap_groups[apg].vaps:
            if not vap.disabled:
                vap_set.add((vap.ssid_prof.ssid, vap.ssid_prof.opmode, vap.fwd_mode))
        if apn in c.ap_names:
            for vap_n in c.ap_names[apn].exclude_vaps:
                vap = c.vap_profs[vap_n]
                k = (vap.ssid_prof.ssid, vap.ssid_prof.opmode, vap.fwd_mode)
                vap_set.discard(k)      # will not raise exception
            for vap in c.ap_names[apn].vaps:
                if not vap.disabled:
                    vap_set.add((vap.ssid_prof.ssid, vap.ssid_prof.opmode, vap.fwd_mode))

        global_vap_set[apn] = vap_set       # list of active VAPs on the AP

        mix_flag = 0        # bridge + tunnel/split-tunnel
        for i in vap_set:
            if i not in vap_count:
                vap_count[i] = 0
            vap_count[i] += 1

            #   detect mixed AP (bridge + tunnel/split-tunnel)
            if mix_flag != 3:
                if i[2] == 'bridge':
                    mix_flag |= 1
                else:
                    mix_flag |= 2
                if mix_flag == 3:
                    global_mixed_ap.add(apn)
                    mixed_ap_count += 1

            if i not in global_vap_count:
                global_vap_count[i] = 0
            global_vap_count[i] += 1

            if i[0] in ('0000docomo', '0001docomo', 'NTT-SPOT'):
                continue
            if i not in coex_vaps_sets:
                coex_vaps_sets[i] = set()
                coex_vaps_count[i] = {}
                controller_sets[i] = set()
            controller_sets[i].add(controller)

            for j in vap_set:
                if i == j:
                    continue
                coex_vaps_sets[i].add(j)
                if j not in coex_vaps_count[i]:
                    coex_vaps_count[i][j] = 0
                coex_vaps_count[i][j] += 1

    #print_debug(f"num_static = {num_static}")





    #
    #   Create one row of the summary table
    #

    row = [c.name, model, version, len(apg_set), len(ap_db_tbl)]
    Ses = ''
    for k in sorted(aptype_count.keys()):
        Ses += f'AP-{k} x {aptype_count[k]}\n'
    row.append(Ses[:-1])

    row.append(len(vap_count))    # SSID 数

    #   List of SSIDs
    Ses = ''
    for i in sorted(vap_count.keys()):  # (ssid, opmode, fwd-mode)
        tmp = f'{i[0]} [{vap_count[i]}]'
        Ses += f"{tmp:28} ({i[1]}/{i[2]})\n"
    row.append(Ses[:-1])

    #   Static Ch APs
    row.append(num_static)

    #   mixed AP count
    row.append(mixed_ap_count)

    #   List of AP specific overrides
    row.append(len(c.ap_names))

    #   stats of AP specifics
    _vaps = _excl = _arm = _radio_disable = _reg = _misc = 0
    _misc_set = set()
    for name, apn in c.ap_names.items():
        if apn.vaps:
            _vaps += 1
        if apn.exclude_vaps:
            _excl += 1
        if (apn.dot11a and apn.dot11a.arm_prof) or (apn.dot11g and apn.dot11g.arm_prof):
            _arm += 1
        if (apn.dot11a and apn.dot11a.name in ('11a-disable', '11g-disable', '11bg-disable', 'no-radio-11g', 'no-radio-11a')) or \
           (apn.dot11g and apn.dot11g.name in ('11a-disable', '11g-disable', '11bg-disable', 'no-radio-11g', 'no-radio-11a')):
            _radio_disable += 1
        if apn.reg:
            _reg += 1
        if apn.misc:
            _misc += 1
            for Ses in apn.misc:
                _misc_set.add(Ses)

    row.extend([_vaps, _excl, _arm, _radio_disable, _reg, _misc, '\n'.join(sorted(_misc_set))])

    # for apn in c.apn_overrides.keys():
    #     print(apn)

    results.append(row)

#
#   Create Excel 1
#
df = pd.DataFrame(results,
                  columns=['Controller', 'Model', 'Version', 'Active groups', 'Active APs', 'AP model', 'Num of SSIDs', 'SSIDs',
                           'Static Ch APs', 'Mixed APs', 'AP specific', 'Add VAP', 'Del VAP', 'ARM', 'Radio OFF', 'Regulatory', 'Misc', 'Misc config'])
wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

f = Font(name='Arial')
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.font = f

widths = [20, 15, 10, 10, 10, 20, 10, 60,       # A-H
          10, 10, 10, 10, 10, 10, 10, 10, 10, 60]   # I-R
for i, w in enumerate(widths):
    ws.column_dimensions[chr(65 + i)].width = w

#   Header format
f = Font(name='Arial', bold=True, size=9)
Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
for cell in ws[1]:
    cell.fill = Ses
    cell.font = f

ws.auto_filter.ref = "A:" + get_column_letter(ws.max_column)
ws.freeze_panes = "A2"

#   multi lines
for c in ws['F']+ws['H']:
    c.alignment = openpyxl.styles.Alignment(wrapText=True)

#   alignment
al = openpyxl.styles.Alignment(vertical='center', wrapText=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = al

#   SSID list font
f = Font(name='Consolas', size=10)
for row in ws['H2':f'H{ws.max_row}']:
    row[0].font = f

#   misc config font
for row in ws['R2':f'R{ws.max_row}']:
    row[0].font = f


#
#   output to file
#
print(f"Writing to {xlsfile1} ... ", end="")
wb.save(xlsfile1)
print("done.")


#############################################################################

#
#   Create Excel 2
#

results = []
for vap in sorted(controller_sets.keys()):
    row = [ vap[0], vap[1], vap[2], global_vap_count[vap], '\n'.join(sorted(controller_sets[vap])) ]

    Ses = ''
    for i in sorted(coex_vaps_sets[vap]):
        tmp = f"{i[0]} [{coex_vaps_count[vap][i]}]"
        Ses += f"{tmp:35} ({i[1]}/{i[2]})\n"
    row.append(Ses[:-1])
    results.append(row)

df = pd.DataFrame(results,
                  columns=['SSID', 'Opmode', 'fwd-mode', 'Num of APs', 'Controllers', 'Co-ex VAPs'])

wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

f = Font(name='Consolas')
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.font = f

widths = [40, 15, 15, 10, 20, 80]
for i, w in enumerate(widths):
    ws.column_dimensions[chr(65 + i)].width = w

#   Header format
f = Font(name='Arial', bold=True, size=9)
Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
for cell in ws[1]:
    cell.fill = Ses
    cell.font = f

ws.auto_filter.ref = "A:" + get_column_letter(ws.max_column)
ws.freeze_panes = "A2"

#   multi lines
for c in ws['E']+ws['F']:
    c.alignment = openpyxl.styles.Alignment(wrapText=True)

#   alignment
al = openpyxl.styles.Alignment(vertical='center', wrapText=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = al

#   VAP list font
# f = Font(name='Consolas', size=10)
# for cell in ws['F2':f'F{ws.max_row}']:
#     cell[0].font = f



#
#   output to file
#
print(f"Writing to {xlsfile2} ... ", end="")
wb.save(xlsfile2)
print("done.")


#############################################################################

#
#   Create Excel 3 (AP name vs VAP)
#

#   sort AP
apn_sorted = sorted(global_apn_set)

vap_list = {}       # key=vap, val=column number

# col = 6     # col1 = AP Name, col2 = AP model, col3 = Num of SSIDs, col4 = mixed, col5 = AP Group
# for apn in apn_sorted:
#     for vap in sorted(global_vap_set[apn.lower()]):
#         if vap in vap_list:
#             continue
#         vap_list[vap] = col
#         col += 1

#   Order VAP by # of APs
vap_sorted = sorted(global_vap_count.keys(), key=lambda x: global_vap_count[x], reverse=True)
col = 6
for vap in vap_sorted:
    vap_list[vap] = col
    col += 1

vap_header = [ f"  {x[0]} ({x[2][0].upper()}) [{global_vap_count[x]}]" for x in vap_list.keys() ]
vap_fwd_mode = [ x[2] for x in vap_list.keys() ]
vap_cols = len(vap_header)

results = []
for apn in apn_sorted:
    model = int(global_ap_model[apn])
    num_vap = len(global_vap_set[apn.lower()])
    mixed = 'Y' if apn.lower() in global_mixed_ap else ''
    row = [ apn, model, num_vap, mixed, global_apn_apg_dic[apn] ] + [None] * vap_cols
    results.append(row)

df = pd.DataFrame(results, columns=['AP Name', 'Model', 'SSIDs', 'Mixed', 'Group'] + vap_header)

wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

f = Font(name='Consolas', size=9)
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.font = f

#   Header format
f = {
    'tunnel': PatternFill(fgColor='77CCFF', fill_type='solid'),
    'split-tunnel': PatternFill(fgColor='66EECC', fill_type='solid'),
    'bridge': PatternFill(fgColor='FFAAAA', fill_type='solid')
}
al = openpyxl.styles.Alignment(textRotation=90, horizontal='center')
for i, cell in enumerate(ws[1]):
    if i < 5:
        continue        # skip left 5 columns
    cell.alignment = al
    cell.fill = f[vap_fwd_mode[i-5]]

#   column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 25
for i in range(vap_cols):
    ws.column_dimensions[get_column_letter(i+6)].width = 2

#   Center alignment for column B, C, D (Model, Num SSIDs, mixed)
al = openpyxl.styles.Alignment(horizontal='center')
for r in ws.iter_rows(min_row=2, min_col=2, max_col=4):
    for cell in r:
        cell.alignment = al

#   freeze
ws.freeze_panes = "F2"

#   mixed AP
f = PatternFill(fgColor='FFCC00', fill_type='solid')
for i, apn in enumerate(apn_sorted):
    if apn.lower() in global_mixed_ap:
        ws['A'+str(i+2)].fill = f

#   fill cells
f = {
    'tunnel': PatternFill(fgColor='55AAEE', fill_type='solid'),
    'split-tunnel': PatternFill(fgColor='44CCAA', fill_type='solid'),
    'bridge': PatternFill(fgColor='FF8888', fill_type='solid')
}
for i, apn in enumerate(apn_sorted):
    for vap in global_vap_set[apn.lower()]:
        col = vap_list[vap]
        cell = ws[get_column_letter(col) + str(i+2)]
        cell.fill = f[vap[2]]
        cell.value = "　"

# filter
ws.auto_filter.ref = "A:" + get_column_letter(ws.max_column)
#ws.auto_filter.ref = ws.dimensions

#
#   output to file
#
print(f"Writing to {xlsfile3} ... ", end="")
wb.save(xlsfile3)
print("done.")
