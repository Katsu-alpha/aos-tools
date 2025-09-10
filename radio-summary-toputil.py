#
#   radio-summary-toputil.py
#
#   show radio-summary をパースし、5GHz radio を channel util 順にソートし、Excel に出力
#   端末数の column は show ap active から取得
#   2025/1/14   tri-radio 対応
#

import sys
import re
import argparse
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from collections import defaultdict
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd


def atoi(s):
    m = re.match(r'(\d+)', s)
    if m:
        return int(m.group(1))
    else:
        return 0

#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Parse show radio-summary and sort the 5GHz radios by channel utilization")
    parser.add_argument('infile', help="Input file(s)", type=str, nargs='+')
    parser.add_argument('--pattern', '-p', help='regex for AP name', type=str, default='.*')
    parser.add_argument('--band', '-b', help='Radio band', type=str, default='5')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)


    #
    #   parse AP tables
    #
    print("Parsing files ... ", end="")
    aos = AOSParser(args.infile, ["show ap radio-summary"], merge=False)
    radio_summary = aos.get_table("show ap radio-summary")
    if radio_summary is None:
        print("show ap radio-summary output not found.")
        sys.exit(-1)

    aos = AOSParser(args.infile, ["show ap active"], merge=True)
    ap_active_tbl = aos.get_table("show ap active")
    if ap_active_tbl is None:
        print("show ap active output not found.")
        sys.exit(-1)
    print("done.")

    #
    #   show ap active から Radio0, Radio1, Radio2 の client 数取得
    #
    idx_r0 = ap_active_tbl[0].index("Radio 0 Band Ch/EIRP/MaxEIRP/Clients")
    idx_r1 = ap_active_tbl[0].index("Radio 1 Band Ch/EIRP/MaxEIRP/Clients")
    idx_r2 = ap_active_tbl[0].index("Radio 2 Band Ch/EIRP/MaxEIRP/Clients")
    apn_ch_sta = defaultdict(lambda: {})

    for row in ap_active_tbl[1:]:
        apn = row[0]

        r0 = row[idx_r0]
        r = re.search(r":(\d+[SE+\-]?)/[\d\.]+/[\d\.]+/(\d+)$", r0)
        if r:
            ch = r.group(1)
            nsta = int(r.group(2))
            apn_ch_sta[apn][ch] = nsta

        r1 = row[idx_r1]
        r = re.search(r":(\d+[SE+\-]?)/[\d\.]+/[\d\.]+/(\d+)$", r1)
        if r:
            ch = r.group(1)
            nsta = int(r.group(2))
            apn_ch_sta[apn][ch] = nsta

        r2 = row[idx_r2]
        r = re.search(r":(\d+[SE+\-]?)/[\d\.]+/[\d\.]+/(\d+)$", r2)
        if r:
            ch = r.group(1)
            nsta = int(r.group(2))
            apn_ch_sta[apn][ch] = nsta

    #
    #   radio-summary の必要な column のみ取り出す
    #
    tbl = []
    ch_ctr = defaultdict(lambda: 0)
    util_sum = defaultdict(lambda: 0)
    for r in radio_summary:
        apn = r[0]
        if not re.search(args.pattern, apn):
            continue

        apg = r[1]
        apt = r[2]
        band = r[4]
        if not band.startswith(args.band):
            continue
        mode = r[5]     # AP:VHT:56
        if mode == 'AM': continue
        m = re.search(":(\d+[SE+\-]?)", mode)   # Channel
        if m:
            ch = m.group(1)
        m = re.search("([0-9\.]+)/", r[6])   # EIRP/MaxEIRP
        if m:
            eirp = float(m.group(1))
        m = re.search("(-\d+)/(\d+)/(\d+)", r[7])     # NF/U/I
        if m:
            nf   = int(m.group(1))
            util = int(m.group(2))
            intf = int(m.group(3))

        ch_ctr[ch] += 1
        util_sum[ch] += util
        if ch in apn_ch_sta[apn]:
            sta = apn_ch_sta[apn][ch]
        else:
            sta = "na"
        row = [ apn, apg, apt, mode, eirp, sta, nf, intf, util ]
        tbl.append(row)

    # sort by Utilization
    tbl.sort(key=lambda x: x[8], reverse=True)

    print("Name                        Group                           Type  Mode          EIRP    Clients  NF    Util")
    print("----                        -----                           ----  ----          ----    -------  ---   ----")
    for r in tbl:
        print(f"{r[0]:28}{r[1]:32}{r[2]:6}{r[3]:14}{r[4]:>4}{r[5]:>7}  {r[6]:>7}{r[8]:>7}")

    print(f"Total APs: {len(apn_ch_sta)}")
    print(f"Total Radios: {len(tbl)}")
    for ch in sorted(ch_ctr.keys(), key=lambda x: atoi(x)):
        avg = util_sum[ch] / ch_ctr[ch]
        print(f"{ch} - {avg:.2f} ({ch_ctr[ch]} Radios)")


    #
    #   Excel 書き出し
    #
    xlsfile = "toputil.xlsx"

    df_toputil = pd.DataFrame(tbl, columns=["AP Name", "Group", "Type", "Mode", "EIRP (dBm)", "Clients", "Noise (dBm)", "Intf (%)", "Util (%)"])

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df_toputil, index=False, header=True):
        ws.append(r)

    # set column width
    widths = [30, 40, 10, 20, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    # set header font and color
    f = Font(name='Calibri', bold=True, size=11, color="FFFFFF")
    p = PatternFill(fgColor="70AD47", fill_type="solid")
    for cell in ws['A1':'I1'][0]:
        cell.font = f
        cell.fill = p

    # set font, bgcolor
    f = Font(name='Consolas')
    p = PatternFill(fgColor="e2efda", fill_type="solid")
    rn = 1
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = f
            if rn&1:
                cell.fill = p
        rn += 1

    # centering for AP type
    for cell in ws['C'][1:]:
        cell.alignment = Alignment(horizontal='center')


    # set autofilter and freeze panes
    ws.auto_filter.ref = "A:I"
    ws.freeze_panes = "A2"


    #
    #   output to file
    #
    print(f"Writing to {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)
