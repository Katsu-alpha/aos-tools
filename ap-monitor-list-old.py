#/usr/bin/python3 -u
#
#   ap2-monitor-list.py
#
#   show ap monitor ap-list の結果をパースし、valid/interfering に分け、curr_snr でソートして出力
#

import sys
import re
import argparse
import pandas as pd
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Join 'show ap database' and 'show ap active' tables and write the result to an MS Excel file")
    parser.add_argument('infile', help="Input file containing 'show ap monitor ap-list' output", type=str, nargs=1)
    #parser.add_argument('outfile', help='Output Excel file', type=str, nargs='?', default='')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)


    #
    #   parse AP List
    #
    print("Parsing files ... ", end="")
    cmd = "show ap monitor ap-list .+"
    aos = AOSParser(args.infile, cmd)
    ap_list_tbl = aos.get_table(cmd)
    if ap_list_tbl is None:
        print("show ap monitor ap-list output not found.")
        sys.exit(-1)
    print("done.")

    #
    #   Process columns in active AP table
    #
    cols = ["bssid", "essid", "chan", "ap-type", "phy-type", "encr", "curr-snr", "curr-rssi"]
    ap_list_tbl_2 = aos.get_table(cmd, *cols)
    for row in ap_list_tbl_2:
        row[6] = int(row[6])
        row[7] = int(row[7])
        r = re.match("(\d+)", row[2])   # channel
        if r:
            row.append(int(r.group(1)))
        else:
            row.append(0)
    #print(aos.table2str(ap_list_tbl_2))
    #sys.exit(0)

    #
    #   table join & sort
    #
    pd.set_option('display.max_columns', 10)
    pd.set_option('display.max_rows', 300)
    pd.set_option('display.width', 1000)
    cols = ["bssid", "essid", "chan", "ap_type", "phy_type", "encr", "curr_snr", "curr_rssi", "pchan"]
    df_ap_list = pd.DataFrame(ap_list_tbl_2, columns=cols)
    df = df_ap_list.sort_values(['curr_snr'], ascending=False)


    # 11a, valid
    df_11a_valid = df.query('ap_type == "valid" and pchan >= 36')
    df_11a_valid.reset_index(drop=True, inplace=True)
    df_11a_valid.index += 1

    # 11a, valid
    df_11a_valid_base = df.query('bssid.str.endswith("0") and ap_type == "valid" and pchan >= 36')
    df_11a_valid_base.reset_index(drop=True, inplace=True)
    df_11a_valid_base.index += 1


    # 11a, interfering
    df_11a_intf = df.query('ap_type != "valid" and pchan >= 36 and curr_snr > 0')
    df_11a_intf.reset_index(drop=True, inplace=True)
    df_11a_intf.index += 1

    # W52, interfering
    df_11a_w52_intf = df.query('ap_type != "valid" and pchan >= 36 and pchan <=48 and curr_snr > 0')
    df_11a_w52_intf.reset_index(drop=True, inplace=True)
    df_11a_w52_intf.index += 1

    # 11g, valid
    #df_11g_valid = df.query('ap_type == "valid" and pchan < 36')
    #df_11g_valid.reset_index(drop=True, inplace=True)
    #df_11g_valid.index += 1

    # 11g, interfering
    #df_11g_intf = df.query('ap_type != "valid" and pchan < 36')
    #df_11g_intf.reset_index(drop=True, inplace=True)
    #df_11g_intf.index += 1

    # print
    print("****************** 11a, valid APs ******************")
    print(df_11a_valid)
    print("\n\n****************** 11a, valid APs (base BSS) ******************")
    print(df_11a_valid_base)
    print("\n\n****************** 11a, interfering APs ******************")
    print(df_11a_intf)
    print("\n\n****************** 11a(W52), interfering APs ******************")
    print(df_11a_w52_intf)

    print(f"\n\n{len(df_11a_w52_intf)} / {len(df_11a_intf)}")
    #print("****************** 11g, valid APs ******************")
    #print(df_11g_valid)
    #print("****************** 11g, interfering APs ******************")
    #print(df_11g_intf)

    sys.exit(0)
    #
    #   Create Excel
    #
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    f = Font(name='Consolas')
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.font = f

    widths = [25, 30, 10, 20, 20, 13, 15, 10, 10, 15, 10, 10]
    for i,w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    f = Font(name='Arial', bold=True, size=9)
    Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws['A1':'L1'][0]:
        cell.fill = Ses
        cell.font = f

    ws.auto_filter.ref = "A:L"
    ws.freeze_panes = "A2"


    #
    #   output to file
    #
    print(f"Writing to {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)
