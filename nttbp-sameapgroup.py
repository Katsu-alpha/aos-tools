#
#   show run summrize
#
#   同一の AP Group を検出
#

import re
import glob
import sys
import argparse
import mylogger as log
from aos_parser import AOSParser
from colorama import Fore, Style
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def print_debug(msg):
    print(Fore.GREEN + msg + Style.RESET_ALL)

def print_info(msg):
    print(Fore.CYAN + msg + Style.RESET_ALL)

def print_warn(msg):
    print(Fore.YELLOW + msg + Style.RESET_ALL)

def print_err(msg):
    print(Fore.RED + msg + Style.RESET_ALL)


class SSID:
    def __init__(self, name):
        self.name = name
        self.ssid = ''
        self.opmode = 'open'
        self.stealth = False
        self.disabled = False

class VAP:
    def __init__(self, name):
        self.name = name
        self.ssid_prof = None
        self.aaa_prof = None
        self.fwd_mode = 'tunnel'
        self.vlan = 1
        self.disabled = False

class APGroup:
    def __init__(self, name):
        self.name = name
        self.vaps = []
        self.dot11a = None
        self.dot11g = None
        self.apsys = 'default'
        self.conf = ''

class APName:
    def __init__(self, name):
        self.name = name
        self.vaps = []
        self.exclude_vaps = []
        self.dot11a = None
        self.dot11g = None
        self.reg = None
        self.apsys = None
        self.misc = []

class ARM:
    def __init__(self, name):
        self.name = name
        self.disabled = False

class DOT11A:
    def __init__(self, name):
        self.name = name
        self.arm_prof = None
        self.radio = True

class DOT11G:
    def __init__(self, name):
        self.name = name
        self.arm_prof = None
        self.radio = True

class APSys:
    def __init__(self, name):
        self.name = name
        self.ses_acl = None

class AOSConfig:
    def __init__(self, name):
        self.name = name        # ホスト名
        self.num_user_roles = 0 # user-role 数
        self.num_acls = 0       # session ACL 数
        self.ssid_profs = {}    # key=プロファイル名, val=SSID obj
        self.vap_profs = {}     # key=プロファイル名, val=VAP obj
        self.ap_groups = {}     # key=AP グループ名, val=APGroup obj
        self.ap_names = {}      # key=AP 名, val=APName obj
        self.arm_profs = {}
        self.dot11a_profs = {}
        self.dot11g_profs = {}
        self.ap_system_profs = {}


def parse_config(name, config):
    C = AOSConfig(name)

    dot11a_def = DOT11A('default')
    C.dot11a_profs['default'] = dot11a_def
    dot11g_def = DOT11G('default')
    C.dot11g_profs['default'] = dot11g_def

    in_cont = False

    for l in config:
        if not in_cont:

            if l.startswith('ip access-list session '):
                C.num_acls += 1  # ip access-list 数カウント
                continue

            if l.startswith('user-role '):
                C.num_user_roles += 1  # user-role 数カウント
                continue

            if l.startswith('ap system-profile '):
                in_cont = 'apsys'
                name = l[18:].replace('"', '')
                apsys = APSys(name)
                C.ap_system_profs[name] = apsys
                continue

            if l.startswith('wlan ssid-profile '):
                in_cont = 'ssid'
                name = l[18:].replace('"', '')
                ssid = SSID(name)
                C.ssid_profs[name] = ssid
                continue

            if l.startswith('wlan virtual-ap '):
                in_cont = 'vap'
                name = l[16:].replace('"', '')
                vap = VAP(name)
                C.vap_profs[name] = vap
                continue

            if l.startswith('ap-group '):
                in_cont = 'apg'
                name = l[9:].replace('"', '')
                apg = APGroup(name)
                C.ap_groups[name.lower()] = apg
                continue

            if l.startswith('ap-name '):
                in_cont = 'apn'
                name = l[8:].replace('"', '').lower()
                apn = APName(name)
                continue

            if l.startswith('rf arm-profile '):
                in_cont = 'arm'
                name = l[15:].replace('"', '')
                arm = ARM(name)
                C.arm_profs[name] = arm
                continue

            if l.startswith('rf dot11a-radio-profile '):
                in_cont = 'dot11a'
                name = l[24:].replace('"', '')
                dot11a = DOT11A(name)
                C.dot11a_profs[name] = dot11a
                continue

            if l.startswith('rf dot11g-radio-profile '):
                in_cont = 'dot11g'
                name = l[24:].replace('"', '')
                dot11g = DOT11G(name)
                C.dot11g_profs[name] = dot11g
                continue

            continue

        #
        #   inside a context
        #

        l = l.strip()
        if l == '!':
            in_cont = False
            continue

        if in_cont == 'apsys':
            if l.startswith('session-acl '):
                n = l[12:].replace('"', '')
                apsys.ses_acl = n
                continue

            r = re.match(r'dns-domain |lms-ip |bkup-lms-ip |bootstrap-threshold |number_ipsec_retries |secondary-master |shell-passwd |bkup-passwords |lms-preemption|ap-console-password |heartbeat-interval ', l)
            if not r:
                #print_warn(f'ap system-profile {name}: {l}')   # 特殊な system-profile 設定
                continue
            continue

        if in_cont == 'ssid':
            if l.startswith('essid '):
                ssid.ssid = l[6:].replace('"', '')
                continue
            if l.startswith('opmode '):
                ssid.opmode = l[7:]
                continue
            if l.startswith('hide-ssid'):
                ssid.stealth = True
                continue
            if l.startswith('no ssid-enable'):
                ssid.disabled = True
                continue
            continue

        if in_cont == 'vap':
            if l.startswith('aaa-profile '):
                vap.aaa_prof = l[12:].replace('"', '')
                continue
            if l.startswith('ssid-profile '):
                n = l[13:].replace('"', '')
                vap.ssid_prof = C.ssid_profs[n]
                if vap.ssid_prof.disabled:
                    vap.disabled = True
                continue
            if l.startswith('vlan '):
                vap.vlan = int(l[5:])
                continue
            if l.startswith('forward-mode '):
                vap.fwd_mode = l[13:]
                continue
            r = re.match(r'allowed-band |broadcast-filter |deny-inter-user-traffic|cellular-handoff-assist', l)
            if not r:
                pass
                #print_warn(f'virtual-ap {name}: {l}')
            continue

        if in_cont == 'apg':
            apg.conf += l + '\n'
            if l.startswith('virtual-ap '):
                n = l[11:].replace('"', '')
                apg.vaps.append(C.vap_profs[n])
                C.ap_groups[name] = apg
                continue
            if l.startswith('dot11a-radio-profile '):
                n = l[21:].replace('"', '')
                apg.dot11a = C.dot11a_profs[n]
                continue
            if l.startswith('dot11g-radio-profile '):
                n = l[21:].replace('"', '')
                apg.dot11g = C.dot11g_profs[n]
                continue
            if l.startswith('ap-system-profile '):
                n = l[18:].replace('"', '')
                apg.apsys = n
                continue
            continue

        if in_cont == 'apn':
            if l.startswith('virtual-ap '):
                n = l[11:].replace('"', '')
                apn.vaps.append(C.vap_profs[n])
                C.ap_names[name] = apn
                continue
            if l.startswith('exclude-virtual-ap '):
                n = l[19:].replace('"', '')
                apn.exclude_vaps.append(n)
                C.ap_names[name] = apn
                continue
            if l.startswith('dot11a-radio-profile '):
                n = l[21:].replace('"', '')
                apn.dot11a = C.dot11a_profs[n]
                continue
            if l.startswith('dot11g-radio-profile '):
                n = l[21:].replace('"', '')
                apn.dot11g = C.dot11g_profs[n]
                continue
            if l.startswith('regulatory-domain-profile '):
                n = l[26:].replace('"', '')
                apn.reg = n
                continue
            if l.startswith('ap-system-profile '):
                n = l[18:].replace('"', '')
                apn.apsys = n
                continue
            #print_warn(f'ap-name {name}: {l}')
            apn.misc.append(l)
            continue

        if in_cont == 'arm':
            if l.startswith('assignment disable'):
                arm.disabled = True
                continue
            continue

        if in_cont == 'dot11a':
            if l.startswith('no radio-enable'):
                dot11a.radio = False
                continue
            if l.startswith('arm-profile '):
                n = l[12:].replace('"', '')
                try:
                    dot11a.arm_prof = C.arm_profs[n]
                except KeyError:
                    pass
                    #print_err(f"In dot11a radio profile '{name}'")
                    #print_err(f"ARM profile '{n}' referenced but not defined.")
                continue
            continue

        if in_cont == 'dot11g':
            if l.startswith('no radio-enable'):
                dot11g.radio = False
                continue
            if l.startswith('arm-profile '):
                n = l[12:].replace('"', '')
                try:
                    dot11g.arm_prof = C.arm_profs[n]
                except KeyError:
                    pass
                    #print_err(f"In dot11g radio profile '{name}'")
                    #print_err(f"ARM profile '{n}' referenced but not defined.")
                continue
            continue

    return C    # return an AOSConfig object

#-------------------------------------------------------------------------

#
#   start
#

parser = argparse.ArgumentParser(
    description="parse show run/show ap database and summarize")
parser.add_argument('files', help="show run/show ap database long outputs", type=str, nargs='*')
parser.add_argument('--debug', help='debug log', action='store_true')
args = parser.parse_args()

if args.debug:
    log.setloglevel(log.LOG_DEBUG)
else:
    log.setloglevel(log.LOG_INFO)

AP_DATABASE_LONG = "show ap database long"

if '*' in args.files[0]:
    args.files = glob.glob(args.files[0])

xlsfile = 'summary-identical-groups.xlsx'

#
#   process each file
#

# apg_apcount = {}
# apg_controllers = {}
# apg_vaps = {}
# vap_count = {}

result = []

for fn in args.files:
    print(f"Processing file: {fn}")

    # Parse 'show ap database long' table
    aos = AOSParser(fn, cmds=[AP_DATABASE_LONG])

    ap_db_tbl = aos.get_table(AP_DATABASE_LONG)
    if ap_db_tbl is None:
        print_info(f"No active AP found in {fn}")
        continue

    ap_db_tbl = ap_db_tbl[1:]   # discard header


    #
    #   Parse running-config, get model, version
    #
    in_showrun = False
    config = []
    with open(fn) as f:
        for line in f:
            if in_showrun:
                if re.match(r'\([\w\-_]+\) #', line):   # AOS6 prompt
                    break
                if re.match(r'\([\w\-_]+\) \*?\[[\w/]+\] #', line):   # AOS8 prompt
                    break
                config.append(line.rstrip())
                continue
            if re.search("show running-config", line):
                in_showrun = True
                continue
            if re.search(r'show configuration committed /mm/mynode', line):
                in_showrun = True
                continue
            r = re.match(r'ArubaOS \(MODEL: (\w+)\), Version ([\d\.]+)', line)
            if r:
                model = r.group(1)
                version = r.group(2)
                continue

    controller = fn[:-4]
    c = parse_config(controller, config)

    apg_conf_dict = {}  # key=apg.conf, val=apg
    same_apg_lists = {}  # key=apg.name, val=list of same apg names
    active_apg_names = set()

    for r in ap_db_tbl:
        apg = r[1]
        apg_l = apg.lower()
        apn = r[0].lower()

        # if apg not in apg_apcount:
        #     apg_apcount[apg] = 0
        # apg_apcount[apg] += 1

        active_apg_names.add(apg_l)

    for apg_n in sorted(active_apg_names):
        apg = c.ap_groups[apg_n]
        if apg.conf in apg_conf_dict:
            orig = apg_conf_dict[apg.conf].name
            print_warn(f"APG '{apg.name}' is identical to '{orig}'")
            if orig not in same_apg_lists:
                same_apg_lists[orig] = []
            same_apg_lists[orig].append(apg.name)
        else:
            apg_conf_dict[apg.conf] = apg

    for orig in same_apg_lists.keys():
        print(f"{controller}: {orig}", end='')
        row = [controller, orig, c.ap_groups[orig.lower()].conf[:-1]]
        for dup in same_apg_lists[orig]:
            print(f"={dup}", end='')
            row[1] += f"\n{dup}"
        print("")
        result.append(row)

#
#   Create Excel
#


df = pd.DataFrame(result, columns=['Controller', 'AP Groups', 'Config'])

wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

f = Font(name='Consolas')
for row in ws.iter_rows(min_row=1):
    for cell in row:
        cell.font = f

widths = [25, 25, 60]
for i, w in enumerate(widths):
    ws.column_dimensions[chr(65 + i)].width = w

#   Header format
f = Font(name='Arial', bold=True, size=9)
Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
for cell in ws[1]:
    cell.fill = Ses
    cell.font = f

ws.auto_filter.ref = "A:" + get_column_letter(ws.max_column)
ws.freeze_panes = "A2"


#   alignment
al = openpyxl.styles.Alignment(vertical='top', wrapText=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = al

#   Column C
f = Font(name='Consolas', size=9)
for r in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in r:
        cell.font = f

#
#   output to file
#
print(f"Writing to {xlsfile} ... ", end="")
wb.save(xlsfile)
print("done.")
