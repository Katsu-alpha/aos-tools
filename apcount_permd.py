#/usr/bin/python3 -u
#
#   ap2count_permd.py
#
#   show switches と show ap database をパースし、MD ごとの AP 数を Excel に保存
#

import sys
import math
import argparse
import pandas as pd
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('infile', help="Input file containing 'show ap database long' and 'show switches' output", type=str, nargs=1)
    parser.add_argument('outfile', help='Output Excel file', type=str, nargs='?', default='')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)

    if args.outfile == '':
        xlsfile = 'ap-sw-table.xlsx'
    else:
        xlsfile = args.outfile

    #
    #   parse AP tables
    #
    print("Parsing files ... ", end="")
    aos = AOSParser(args.infile, [AP_DATABASE_LONG_TABLE, "show switches"])
    ap_database_tbl = aos.get_table(AP_DATABASE_LONG_TABLE)
    switches_tbl    = aos.get_table("show switches", "IP Address", "Name", "Location", "Model")
    if ap_database_tbl is None:
        print("show ap database long output not found.")
        sys.exit(-1)
    if switches_tbl is None:
        print("show switches output not found.")
        sys.exit(-1)
    print("done.")

    print(f"{len(ap_database_tbl)-1} active APs found.")
    #
    #   count APs per each MD
    #

    sw_ctr = {}
    for sw_ip in aos.get_table(AP_DATABASE_LONG_TABLE, "Switch IP"):
        if sw_ip in sw_ctr:
            sw_ctr[sw_ip] += 1
        else:
            sw_ctr[sw_ip] = 1

    ap_cnt_tbl = [[sw_ip, ctr] for sw_ip, ctr in sw_ctr.items()]

    #
    #   table join & sort
    #

    df_switches = pd.DataFrame(switches_tbl, columns=["IP Address", "Name", "Location", "Model"])
    df_ap_ctr   = pd.DataFrame(ap_cnt_tbl, columns=["IP Address", "Num APs"])

    df1 = df_switches.merge(df_ap_ctr, how="left", left_on="IP Address", right_on="IP Address")
    df = df1.sort_values(['Location', 'Name'])

    #
    #   Count AP per location
    #
    loc_tbl = df[['Location', 'Num APs']].values.tolist()

    loc_ap_cnt = {}
    loc_md_cnt = {}
    for r in loc_tbl:
        if math.isnan(r[1]):
            r[1] = 0
        if r[0] in loc_ap_cnt:
            loc_ap_cnt[r[0]] += r[1]
            loc_md_cnt[r[0]] += 1
        else:
            loc_ap_cnt[r[0]] = r[1]
            loc_md_cnt[r[0]] = 1

    for loc in loc_ap_cnt:
        print(f"{loc}, {int(loc_md_cnt[loc])}, {int(loc_ap_cnt[loc])}")

    sys.exit(0)


    #
    #   Create Excel
    #
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    f = Font(name='Calibri')
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.font = f

    widths = [20, 30, 50, 15, 10]
    for i,w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    f = Font(name='Arial', bold=True, size=9)
    s = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws['A1':'E1'][0]:
        cell.fill = s
        cell.font = f

    ws.auto_filter.ref = "A:E"
    ws.freeze_panes = "A2"


    #
    #   output to file
    #
    print(f"Writing to {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)
