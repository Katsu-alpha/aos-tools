#/usr/bin/python3 -u
#
#   ap2-monitor-list-iap.py
#
#   IAP の show ap monitor ap-list の結果をパースし、valid/interfering に分け、curr_snr でソートして出力
#

import sys
import re
import argparse
import pandas as pd
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pprint as pp

#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Parse show ap monitor ap-list and summarize to an excel file")
    parser.add_argument('infile', help="Input file containing 'show ap monitor ap-list' output", type=str, nargs=1)
    #parser.add_argument('outfile', help='Output Excel file', type=str, nargs='?', default='')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)


    #
    #   parse AP List
    #
    print("Parsing files ... ", end="")
    cmd = "show ap monitor ap-list"
    aos = AOSParser(args.infile, cmd)
    ap_list_tbl = aos.get_table(cmd)
    if ap_list_tbl is None:
        print("show ap monitor ap-list output not found.")
        sys.exit(-1)
    print("done.")

    #
    #   Process columns in active AP table
    #
    cols = ["bssid", "essid", "band/chan/ch-width/ht-type", "ap-type", "encr", "curr-snr", "curr-rssi"]
    ap_list_tbl_2 = aos.get_table(cmd, *cols)
    # pp.pprint(ap_list_tbl_2)
    # exit()
    data = []
    for row in ap_list_tbl_2:
        band, chan, cw, ht_type = row[2].split('/')
        curr_snr = int(row[5])
        curr_rssi = int(row[6])
        r = re.match("(\d+)", chan)   # get primary channel
        if r:
            pchan = int(r.group(1))
        else:
            pchan = 0
        data.append([row[0], row[1], band, chan, ht_type, row[3], row[4], curr_snr, curr_rssi])

    cols = ["bssid", "essid", "band", "chan", "ht_type", "ap_type", "encr", "curr_snr", "curr_rssi"]
    df_ap_list = pd.DataFrame(data, columns=cols)
    df = df_ap_list.sort_values(['band', 'curr_snr'], ascending=False)

    #
    #   Create Excel
    #
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    f = Font(name='Consolas')
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = f

    # "bssid", "essid", "band", "chan", "ht_type", "ap_type", "encr", "curr_snr", "curr_rssi"
    widths = [25, 30, 10, 10, 20, 20, 15, 10, 10]
    for i,w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    # Header font and color
    f = Font(name='Arial', bold=True, size=9)
    s = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws['A1':'I1'][0]:
        cell.fill = s
        cell.font = f

    ws.auto_filter.ref = "A:I"
    ws.freeze_panes = "A2"


    #
    #   output to file
    #
    xlsfile = "ap-list.xlsx"
    print(f"Writing to {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)
