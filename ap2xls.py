#/usr/bin/python3 -u
#
#   ap2xls.py
#
#   Convert AP database table to MS Excel
#
#   2024/10/24 - AP-655 (tri-radio) 対応

import sys
import re
import argparse
import pandas as pd
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict

def uniq(tbl):
    ret = []
    k = set()
    for r in tbl:
        if r[0] in k: continue
        k.add(r[0])
        ret.append(r)
    return ret

def toi(s):
    try:
        return int(s)
    except ValueError:
        return 0


#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Convert AP database table to MS Excel")
    parser.add_argument('infile', help="Input file(s) containing 'show ap database long' and 'show ap active' output", type=str, nargs='+')
    parser.add_argument('outfile', help='Output Excel file', type=str, nargs='?', default='')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    parser.add_argument('--pattern', '-p', help='regex for AP name', type=str, default='.*')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)

    if args.outfile == '':
        xlsfile = 'ap-table.xlsx'
    else:
        xlsfile = args.outfile

    #
    #   parse AP tables
    #
    print("Parsing files ... ", end="")
    try:
        aos = AOSParser(args.infile, [AP_DATABASE_LONG_TABLE], merge=True, encoding='utf-8')
    except UnicodeDecodeError as e:
        print(f'UTF-8 decode error. Trying Shift-JIS...')
        aos = AOSParser(args.infile, [AP_DATABASE_LONG_TABLE], merge=True, encoding='shift-jis')

    ap_database_tbl = aos.get_table(AP_DATABASE_LONG_TABLE)
    if ap_database_tbl is None:
        print("show ap database long output not found.")
        sys.exit(-1)
    print("done.")

    ap_database_tbl = uniq(ap_database_tbl)
    print(f"{len(ap_database_tbl)-1} unique APs found in ap database.")

    #
    #   filter AP database table
    #
    ap_db_tbl = [ap_database_tbl[0]]
    for r in ap_database_tbl[1:]:
        if not re.search(args.pattern, r[0]):
            continue
        ap_db_tbl.append(r)

    #
    #   AP model tally
    #
    apmodelctr = defaultdict(lambda: 0)
    for r in ap_db_tbl[1:]:
        model = r[2]
        apmodelctr[model] += 1
    print("AP models")
    for model in sorted(apmodelctr.keys()):
        print(f"{model}: {apmodelctr[model]}")
    print()



    #
    #   table join & sort
    #
    df_ap_database = pd.DataFrame(ap_db_tbl[1:], columns=ap_db_tbl[0])
    df = df_ap_database[[
        'Name', 'Group', 'AP Type', 'IP Address', 'Switch IP',
        'Serial #', 'Wired MAC Address',
    ]]
    #df = df.sort_values(['Group', 'Name'])
    df = df.sort_values(['Name'])

    #
    #   Create Excel
    #
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    f = Font(name='Consolas')
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.font = f

    widths = [25, 30, 10, 20, 20, 13, 25]
    for i,w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    f = Font(name='Arial', bold=True, size=9)
    Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws['A1':'G1'][0]:
        cell.fill = Ses
        cell.font = f

    ws.auto_filter.ref = "A:G"
    ws.freeze_panes = "A2"


    #
    #   output to file
    #
    print(f"Writing to {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)
