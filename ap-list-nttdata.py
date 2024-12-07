#/usr/bin/python3 -u
#
#   AP list parser for NTT-Data
#
#   show ap monitor ap-list の結果をパースし、valid/interfering に分け、curr_snr でソートして出力
#

import sys
import re
import argparse
import pandas as pd
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Join 'show ap database' and 'show ap active' tables and write the result to an MS Excel file")
    parser.add_argument('infiles', help="Input file containing 'show ap monitor ap-list' output", type=str, nargs='+')
    #parser.add_argument('outfile', help='Output Excel file', type=str, nargs='?', default='')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)


    #
    #   parse BSSID List
    #
    cmd = "show ap bss-table"
    aos = AOSParser('bss-table.txt', cmd)
    bss_table = aos.get_tables(cmd, 'bss', 'ap name')
    if bss_table is None:
        print("show ap bss-table output not found.")
        sys.exit(-1)

    bss2apn = {}
    for tbl in bss_table:
        for r in tbl:
            bss2apn[r[0]] = r[1]

    print(f"{len(bss2apn)} BSSIDs found.")


    #
    #   parse AP List
    #
    for fn in args.infiles:
        xlsfile = fn[:-4] + '.xlsx'
        print("Parsing files ... ", end="")
        cmd = "show ap monitor ap-list .+"
        aos = AOSParser(fn, cmd)
        ap_list_tbl = aos.get_table(cmd)
        if ap_list_tbl is None:
            print("show ap monitor ap-list output not found.")
            continue
        print("done.")

        #
        #   Process columns in ap-list table
        #
        cols = ["bssid", "essid", "chan", "ap-type", "phy-type", "encr", "curr-snr", "curr-rssi"]
        ap_list_tbl_2 = aos.get_table(cmd, *cols)
        new_tbl = []
        for row in ap_list_tbl_2:
            if not row[4].startswith('80211a'):
                continue
            if row[3] != 'valid':
                continue
            snr  = int(row[6])
            rssi = -int(row[7])
            r = re.match("(\d+)", row[2])   # channel
            if r:
                pchan = int(r.group(1))
            else:
                pchan = 0
            bss = row[0]
            apn = ''
            if bss in bss2apn:
                apn = bss2apn[bss]
            row.insert(2, apn)
            new_tbl.append([row[0], row[1], apn, pchan, snr, rssi])

        #
        #   table join & sort
        #
        cols = ["BSSID", "ESSID", "AP Name", "Chan", "SNR", "RSSI"]
        df_ap_list = pd.DataFrame(new_tbl, columns=cols)
        df = df_ap_list.sort_values(['SNR'], ascending=False)

        #
        #   Create Excel
        #
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        f = Font(name='Consolas')
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.font = f

        # "bssid", "essid", "AP Name", "chan", "SNR", "RSSI"
        widths = [25, 30, 20, 10, 10, 10]
        for i,w in enumerate(widths):
            ws.column_dimensions[chr(65+i)].width = w

        f = Font(name='Arial', bold=True, size=9)
        Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
        for cell in ws['A1':'F1'][0]:
            cell.fill = Ses
            cell.font = f

        ws.auto_filter.ref = "A:F"
        ws.freeze_panes = "A2"


        #
        #   output to file
        #
        print(f"Writing to {xlsfile} ... ", end="")
        wb.save(xlsfile)
        print("done.")

    sys.exit(0)
