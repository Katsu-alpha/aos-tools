#/usr/bin/python3 -u
#
#   ap2xls_denso.py
#
#   Join 'show ap database' and 'show ap active' tables and write the result to an MS Excel file
#   add following columns
#       AP prefix, Radio 0 PHY, Radio 0 Ch, Radio 0 EIRP
#

import sys
import argparse
import re
import pandas as pd
import mylogger as log
from aos_parser import AOSParser, AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


#
#   main
#
if __name__ == '__main__':

    parser = argparse.ArgumentParser(
        description="Join 'show ap database' and 'show ap active' tables and write the result to an MS Excel file")
    parser.add_argument('infile', help="Input file containing 'show ap database long' and 'show ap active' output", type=str, nargs=1)
    parser.add_argument('outfile', help='Output Excel file', type=str, nargs='?', default='')
    parser.add_argument('--debug', help='Enable debug log', action='store_true')
    args = parser.parse_args()

    if args.debug:
        log.setloglevel(log.LOG_DEBUG)
    else:
        log.setloglevel(log.LOG_INFO)

    if args.outfile == '':
        xlsfile = 'ap-table.xlsx'
    else:
        xlsfile = args.outfile

    #
    #   parse AP tables
    #
    print("Parsing files ... ", end="")
    aos = AOSParser(args.infile, [AP_DATABASE_LONG_TABLE, AP_ACTIVE_TABLE])
    ap_database_tbl = aos.get_table(AP_DATABASE_LONG_TABLE)
    ap_active_tbl   = aos.get_table(AP_ACTIVE_TABLE)
    if ap_database_tbl is None:
        print("show ap database long output not found.")
        sys.exit(-1)
    if ap_active_tbl is None:
        print("show ap active output not found.")
        sys.exit(-1)
    print("done.")


    #
    #   Process columns in active AP table
    #
    ap_act_tbl = []
    ap_act_tbl.append(["Name", "Prefix", "Radio 0 PHY", "Radio 0 Ch", "Radio 0 EIRP", "Radio 1 PHY", "Radio 1 Ch", "Radio 1 EIRP"])
    idx_r0 = ap_active_tbl[0].index("Radio 0 Band Ch/EIRP/MaxEIRP/Clients")
    idx_r1 = ap_active_tbl[0].index("Radio 1 Band Ch/EIRP/MaxEIRP/Clients")
    idx_name = ap_active_tbl[0].index("Name")

    for row in ap_active_tbl[1:]:
        r0 = row[idx_r0]
        r = re.match("(.+):([\dSE+\-]+)/([\d\.]+)/[\d\.]+/\d+$", r0)
        if r:
            r0_phy  = r.group(1)
            r0_ch   = r.group(2)
            r0_eirp = r.group(3)
        else:
            r0_phy = ""
            r0_ch = ""
            r0_eirp = ""

        r1 = row[idx_r1]
        r = re.match("(.+):([\dSE+\-]+)/([\d\.]+)/[\d\.]+/\d+$", r1)
        if r:
            r1_phy  = r.group(1)
            r1_ch   = r.group(2)
            r1_eirp = r.group(3)
        else:
            r1_phy = ""
            r1_ch = ""
            r1_eirp = ""

        name = row[idx_name]
        r = re.match("([^-]+-[^-]+-[^-]+)-", name)
        if r:
            pref = r.group(1)
        else:
            pref = ""

        ap_act_tbl.append([row[idx_name], pref, r0_phy, r0_ch, r0_eirp, r1_phy, r1_ch, r1_eirp])


    #
    #   table join & sort
    #
    df_ap_database = pd.DataFrame(ap_database_tbl[1:], columns=ap_database_tbl[0])
    df_ap_act      = pd.DataFrame(ap_act_tbl[1:], columns=ap_act_tbl[0])

    df1 = df_ap_database.merge(df_ap_act, how="left", left_on="Name", right_on="Name")
    df2 = df1[[
        'Name', 'Prefix', 'Group', 'AP Type', 'IP Address', 'Switch IP', 'Serial #',
        'Radio 0 PHY', 'Radio 0 Ch', 'Radio 0 EIRP',
        'Radio 1 PHY', 'Radio 1 Ch', 'Radio 1 EIRP',
    ]]
    df = df2.sort_values(['Group', 'Name'])


    #
    #   Create Excel
    #
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    f = Font(name='Consolas')
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.font = f

    widths = [25, 14, 20, 10, 20, 20, 13, 15, 6, 6, 15, 6, 6]
    for i,w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    f = Font(name='Arial', bold=True, size=9)
    Ses = PatternFill(fgColor="BDD7EE", fill_type="solid")
    for cell in ws['A1':'M1'][0]:
        cell.fill = Ses
        cell.font = f

    ws.auto_filter.ref = "A:M"
    ws.freeze_panes = "A2"


    #
    #   output to file
    #
    print(f"Writing to {xlsfile} ... ", end="")
    wb.save(xlsfile)
    print("done.")

    sys.exit(0)
